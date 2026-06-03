import json
import os
from decimal import Decimal
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.db.models import Sum, Count, Q, F, Avg
from django.utils import timezone
from projets.models import Projet, Secteur, Programme, TAUX_VERS_FCFA
from bailleurs.models import Bailleur
from financements.models import Financement, Decaissement
from pnd.models import PlanNational, Pilier, SousObjectif
from accounts.decorators import login_required_custom


def _decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError


def _get_user_bailleur_ids(user):
    """Retourne la liste des IDs bailleurs visibles, ou None pour tout voir."""
    if user.is_superuser:
        return None
    profile = getattr(user, 'profile', None)
    if not profile:
        return []
    return profile.get_visible_bailleur_ids()


@login_required_custom
def index(request):
    # Redirige Ministre vers son tableau de bord synthétique dédié
    profile = getattr(request.user, 'profile', None)
    if profile and profile.fonction == 'ministre' and not request.user.is_superuser:
        from django.shortcuts import redirect
        return redirect('dashboard:ministre')
    today = timezone.now().date()
    now = timezone.now()
    one_month_ago = now - timezone.timedelta(days=30)

    # ── Filtrage par bailleurs autorisés ──
    bailleur_ids = _get_user_bailleur_ids(request.user)

    if bailleur_ids is None:
        projets_qs = Projet.objects.all()
        bailleurs_qs = Bailleur.objects.all()
        financements_qs = Financement.objects.all()
        decaissements_qs = Decaissement.objects.all()
    else:
        bailleurs_qs = Bailleur.objects.filter(pk__in=bailleur_ids)
        projets_qs = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()
        financements_qs = Financement.objects.filter(bailleur_id__in=bailleur_ids)
        decaissements_qs = Decaissement.objects.filter(financement__bailleur_id__in=bailleur_ids)

    # ── KPIs ──
    total_projets = projets_qs.count()
    total_programmes = Programme.objects.count()
    projets_en_cours = projets_qs.filter(statut='en_cours').count()
    # Comptage en retard via critères configurables
    try:
        from alertes.models import CritereRetard
        _criteres = list(CritereRetard.objects.filter(actif=True))
    except Exception:
        _criteres = []
    if _criteres:
        _en_cours = list(projets_qs.filter(statut='en_cours').only(
            'pk', 'taux_avancement', 'montant_total', 'devise',
            'taux_decaissement_prevu_annee', 'date_debut', 'date_fin_prevue',
        ))
        projets_en_retard = sum(1 for p in _en_cours if any(c.evaluer(p) for c in _criteres))
    else:
        projets_en_retard = projets_qs.filter(
            statut='en_cours', date_fin_prevue__lt=today
        ).count()
    total_bailleurs = bailleurs_qs.count()
    taux_avancement_global = round(float(
        projets_qs.filter(statut='en_cours').aggregate(avg=Avg('taux_avancement'))['avg'] or 0
    ), 1)

    total_engage = financements_qs.aggregate(total=Sum('montant_engage'))['total'] or 0
    total_montant_projets = projets_qs.aggregate(total=Sum('montant_total'))['total'] or 0
    total_decaisse = decaissements_qs.aggregate(total=Sum('montant'))['total'] or 0

    # ── Agrégats en FCFA (toutes devises converties) ──
    def _to_fcfa(montant, devise):
        return float(montant) * float(TAUX_VERS_FCFA.get(devise, 1))

    total_montant_fcfa = sum(
        _to_fcfa(p['montant_total'], p['devise'])
        for p in projets_qs.values('montant_total', 'devise')
    )
    total_engage_fcfa = sum(
        _to_fcfa(f['montant_engage'], f['devise'])
        for f in financements_qs.values('montant_engage', 'devise')
    )
    total_decaisse_fcfa = sum(
        _to_fcfa(d.montant, d.financement.devise)
        for d in decaissements_qs.select_related('financement')
    )
    taux_decaissement_global = round(
        total_decaisse_fcfa / total_engage_fcfa * 100, 1
    ) if total_engage_fcfa > 0 else 0

    # ── Objectif de décaissement pour l'année en cours (en FCFA) ──
    annee_courante = today.year
    decaisse_annee_fcfa = sum(
        _to_fcfa(d.montant, d.financement.devise)
        for d in decaissements_qs.filter(date_decaissement__year=annee_courante)
        .select_related('financement')
    )
    projets_avec_objectif = projets_qs.filter(
        statut='en_cours', taux_decaissement_prevu_annee__gt=0
    ).only('montant_total', 'devise', 'taux_decaissement_prevu_annee')
    total_prevu_annee_fcfa = sum(
        _to_fcfa(p.montant_total, p.devise) * float(p.taux_decaissement_prevu_annee) / 100
        for p in projets_avec_objectif
    )
    taux_realisation_annuel = round(
        decaisse_annee_fcfa / total_prevu_annee_fcfa * 100, 1
    ) if total_prevu_annee_fcfa > 0 else None
    nb_projets_avec_objectif = projets_avec_objectif.count()

    # ── Variation vs last month ──
    projets_prev = projets_qs.filter(date_creation__lt=one_month_ago).count()
    var_projets = total_projets - projets_prev

    montant_prev = float(projets_qs.filter(
        date_creation__lt=one_month_ago
    ).aggregate(t=Sum('montant_total'))['t'] or 0)
    var_montant = float(total_montant_projets) - montant_prev

    engage_prev = float(financements_qs.filter(
        date_creation__lt=one_month_ago
    ).aggregate(t=Sum('montant_engage'))['t'] or 0)
    var_engage = float(total_engage) - engage_prev

    decaisse_prev = float(decaissements_qs.filter(
        date_decaissement__lt=one_month_ago
    ).aggregate(t=Sum('montant'))['t'] or 0)
    var_decaisse = float(total_decaisse) - decaisse_prev

    retard_prev = projets_qs.filter(
        statut='en_cours', date_fin_prevue__lt=one_month_ago
    ).count()
    var_retard = projets_en_retard - retard_prev  # approximation acceptable pour la variation

    # ── Raw datasets for client-side analytics engine ──

    # Bailleurs with category info
    bailleurs_list = list(
        bailleurs_qs.values('id', 'nom', 'sigle', 'type_bailleur', 'categorie_institutionnelle', 'pays_siege')
    )
    # Enrich with financials
    for b in bailleurs_list:
        b['label'] = b['sigle'] or b['nom'][:20]
        b['engage'] = float(financements_qs.filter(bailleur_id=b['id']).aggregate(t=Sum('montant_engage'))['t'] or 0)
        b['decaisse'] = float(decaissements_qs.filter(financement__bailleur_id=b['id']).aggregate(t=Sum('montant'))['t'] or 0)
        b['nb_projets'] = financements_qs.filter(bailleur_id=b['id']).values('projet').distinct().count()
        cat_map = dict(Bailleur.CATEGORIE_CHOICES)
        b['categorie_label'] = cat_map.get(b['categorie_institutionnelle'], 'Autre')

    # All projects with denormalized fields for analytics
    projets_list = []
    for p in projets_qs.select_related('secteur', 'bailleur_principal').all():
        projets_list.append({
            'id': p.id,
            'code': p.code,
            'titre': p.titre[:60],
            'secteur': p.secteur.nom if p.secteur else 'Non défini',
            'secteur_couleur': p.secteur.couleur if p.secteur else '#94A3B8',
            'bailleur_id': p.bailleur_principal_id,
            'bailleur': (p.bailleur_principal.sigle or p.bailleur_principal.nom[:20]) if p.bailleur_principal else 'Non défini',
            'bailleur_categorie': p.bailleur_principal.categorie_institutionnelle if p.bailleur_principal else 'autre',
            'est_cofinance': p.est_cofinance,
            'nombre_bailleurs': p.nombre_bailleurs,
            'bailleurs_list': [{'id': b.id, 'sigle': b.sigle or b.nom[:20]} for b in p.bailleurs_list],
            'statut': p.get_statut_display(),
            'statut_code': p.statut,
            'montant': float(p.montant_total),
            'devise': p.devise,
            'zone': p.zone_geographique or 'Non précisé',
            'taux_avancement': float(p.taux_avancement),
            'taux_decaissement': float(p.taux_decaissement),
            'date_signature': p.date_signature.isoformat() if p.date_signature else None,
            'date_debut': p.date_debut.isoformat() if p.date_debut else None,
            'date_fin_prevue': p.date_fin_prevue.isoformat() if p.date_fin_prevue else None,
            'en_retard': p.est_en_retard,
        })

    # Financements (filtered)
    financements_list = list(
        financements_qs.select_related('projet', 'bailleur').values(
            'id', 'projet__code', 'projet__titre',
            'bailleur__sigle', 'bailleur__nom', 'bailleur__categorie_institutionnelle',
            'bailleur_id',
            'type_financement', 'montant_engage', 'devise',
        )
    )
    for f in financements_list:
        f['montant_engage'] = float(f['montant_engage'])
        f['bailleur_label'] = f['bailleur__sigle'] or (f['bailleur__nom'] or '')[:20]
        dec_total = Decaissement.objects.filter(financement_id=f['id']).aggregate(t=Sum('montant'))['t'] or 0
        f['decaisse'] = float(dec_total)

    # Secteurs
    secteurs_list = list(
        Secteur.objects.annotate(nb_projets=Count('projet')).values('id', 'nom', 'couleur', 'nb_projets')
    )

    # Statut choices for filters
    statut_choices = [{'code': code, 'label': label} for code, label in Projet.STATUT_CHOICES]

    # Category choices
    categorie_choices = [{'code': code, 'label': label} for code, label in Bailleur.CATEGORIE_CHOICES]

    # Zones (unique)
    zones = sorted(set(p['zone'] for p in projets_list if p['zone'] != 'Non précisé'))

    # ── Couverture PND ──
    plan_actif = PlanNational.objects.filter(actif=True).first()
    piliers_data = []
    if plan_actif:
        for pilier in plan_actif.piliers.all():
            nb = Projet.objects.filter(objectifs_pnd__pilier=pilier).distinct().count()
            montant = Financement.objects.filter(
                projet__objectifs_pnd__pilier=pilier
            ).distinct().aggregate(total=Sum('montant_engage'))['total'] or 0
            piliers_data.append({
                'pilier': pilier,
                'nb_projets': nb,
                'montant': float(montant),
            })

    # ── Lists for tables ──
    derniers_projets = projets_qs.select_related(
        'secteur', 'bailleur_principal'
    ).order_by('-date_creation')[:5]

    if _criteres:
        _candidats = list(
            projets_qs.filter(statut='en_cours')
            .select_related('secteur', 'bailleur_principal')
            .order_by('date_fin_prevue')
        )
        _retard_items = []
        for p in _candidats:
            motifs = [c.get_label_declenchement(p) for c in _criteres if c.evaluer(p)]
            if motifs:
                p.motifs_retard = motifs
                _retard_items.append(p)
                if len(_retard_items) >= 5:
                    break
        projets_retard_list = _retard_items
    else:
        _base = list(
            projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today)
            .select_related('secteur', 'bailleur_principal')
            .order_by('date_fin_prevue')[:5]
        )
        for p in _base:
            p.motifs_retard = ['Date de fin dépassée']
        projets_retard_list = _base

    # ── Regions of Côte d'Ivoire for map ──
    ci_regions = [
        {'nom': 'Abidjan', 'lat': 5.3600, 'lng': -4.0083},
        {'nom': 'Yamoussoukro', 'lat': 6.8276, 'lng': -5.2893},
        {'nom': 'Bouaké', 'lat': 7.6881, 'lng': -5.0305},
        {'nom': 'San-Pédro', 'lat': 4.7392, 'lng': -6.6363},
        {'nom': 'Daloa', 'lat': 6.8774, 'lng': -6.4502},
        {'nom': 'Korhogo', 'lat': 9.4580, 'lng': -5.6292},
        {'nom': 'Man', 'lat': 7.4127, 'lng': -7.5539},
        {'nom': 'Gagnoa', 'lat': 6.1319, 'lng': -5.9506},
        {'nom': 'Odienné', 'lat': 9.5085, 'lng': -7.5660},
        {'nom': 'Bondoukou', 'lat': 8.0400, 'lng': -2.8000},
        {'nom': 'Sassandra', 'lat': 4.9500, 'lng': -6.0833},
        {'nom': 'Divo', 'lat': 5.8372, 'lng': -5.3571},
        {'nom': 'Abengourou', 'lat': 6.7297, 'lng': -3.4964},
        {'nom': 'Agboville', 'lat': 5.9282, 'lng': -4.2132},
        {'nom': 'Séguéla', 'lat': 7.9614, 'lng': -6.6731},
        {'nom': 'Dabou', 'lat': 5.3256, 'lng': -4.3767},
        {'nom': 'Grand-Bassam', 'lat': 5.2139, 'lng': -3.7340},
        {'nom': 'Ferkessédougou', 'lat': 9.5935, 'lng': -5.1986},
        {'nom': 'Dimbokro', 'lat': 6.6500, 'lng': -4.7000},
        {'nom': 'Bouaflé', 'lat': 6.9833, 'lng': -5.7500},
        {'nom': 'Issia', 'lat': 6.4900, 'lng': -6.5800},
        {'nom': 'Katiola', 'lat': 8.1400, 'lng': -5.1000},
        {'nom': 'Soubré', 'lat': 5.7833, 'lng': -6.5833},
        {'nom': 'Tingréla', 'lat': 10.4833, 'lng': -6.3833},
        {'nom': 'National', 'lat': 7.54, 'lng': -5.55},
    ]

    context = {
        'total_projets': total_projets,
        'total_programmes': total_programmes,
        'projets_en_cours': projets_en_cours,
        'taux_avancement_global': taux_avancement_global,
        'projets_en_retard': projets_en_retard,
        'total_bailleurs': total_bailleurs,
        'total_engage': total_engage,
        'total_montant_projets': total_montant_projets,
        'total_decaisse': total_decaisse,
        'taux_decaissement_global': json.dumps(taux_decaissement_global),
        'total_montant_fcfa': total_montant_fcfa,
        'total_engage_fcfa': total_engage_fcfa,
        'total_decaisse_fcfa': total_decaisse_fcfa,
        'decaisse_annee': decaisse_annee_fcfa,
        'total_prevu_annee': total_prevu_annee_fcfa,
        'taux_realisation_annuel': taux_realisation_annuel,
        'nb_projets_avec_objectif': nb_projets_avec_objectif,
        'annee_courante': annee_courante,
        'var_projets': var_projets,
        'var_montant': var_montant,
        'var_engage': var_engage,
        'var_decaisse': var_decaisse,
        'var_retard': var_retard,
        'piliers_data': piliers_data,
        'plan_actif': plan_actif,
        'derniers_projets': derniers_projets,
        'projets_retard_list': projets_retard_list,
        # JSON datasets for analytics engine
        'projets_json': json.dumps(projets_list, default=_decimal_default),
        'bailleurs_json': json.dumps(bailleurs_list, default=_decimal_default),
        'financements_json': json.dumps(financements_list, default=_decimal_default),
        'secteurs_json': json.dumps(secteurs_list, default=_decimal_default),
        'statut_choices_json': json.dumps(statut_choices),
        'categorie_choices_json': json.dumps(categorie_choices),
        'zones_json': json.dumps(zones),
        'ci_regions_json': json.dumps(ci_regions),
    }
    return render(request, 'dashboard/index.html', context)


@login_required_custom
def regions_geojson(request):
    filepath = os.path.join(settings.BASE_DIR, 'static', 'data', 'regions.geojson')
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    response = JsonResponse(data, json_dumps_params={'ensure_ascii': False})
    response['Cache-Control'] = 'public, max-age=86400'
    return response


@login_required_custom
def api_search(request):
    """Global search across projets, bailleurs, financements."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    results = []

    # Search projets
    for p in Projet.objects.filter(
        Q(code__icontains=q) | Q(titre__icontains=q)
    ).select_related('bailleur_principal', 'secteur')[:8]:
        results.append({
            'type': 'projet',
            'icon': 'folder_open',
            'label': f'[{p.code}] {p.titre[:60]}',
            'sub': p.bailleur_principal.sigle if p.bailleur_principal else '',
            'url': f'/projets/{p.pk}/',
        })

    # Search bailleurs
    for b in Bailleur.objects.filter(
        Q(nom__icontains=q) | Q(sigle__icontains=q)
    )[:5]:
        results.append({
            'type': 'bailleur',
            'icon': 'account_balance',
            'label': b.nom,
            'sub': b.get_type_bailleur_display(),
            'url': f'/bailleurs/{b.pk}/',
        })

    # Search financements by project code or bailleur
    for f in Financement.objects.filter(
        Q(projet__code__icontains=q) | Q(bailleur__sigle__icontains=q) | Q(reference__icontains=q)
    ).select_related('projet', 'bailleur')[:5]:
        results.append({
            'type': 'financement',
            'icon': 'payments',
            'label': f'{f.bailleur.sigle or f.bailleur.nom[:20]} → {f.projet.code}',
            'sub': f.get_type_financement_display(),
            'url': f'/financements/{f.pk}/',
        })

    return JsonResponse({'results': results[:15]})


@login_required_custom
def api_notifications(request):
    """Return actionable notifications/alerts filtered by user's bailleurs."""
    today = timezone.now().date()
    notifs = []

    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets_qs = Projet.objects.all()
    elif not bailleur_ids:
        return JsonResponse({'notifications': [], 'count': 0})
    else:
        projets_qs = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()

    # Projects en retard
    retards = projets_qs.filter(
        statut='en_cours', date_fin_prevue__lt=today
    ).select_related('bailleur_principal').order_by('date_fin_prevue')[:5]
    for p in retards:
        days_late = (today - p.date_fin_prevue).days
        notifs.append({
            'type': 'warning',
            'icon': 'warning',
            'title': f'{p.code} en retard',
            'message': f'{days_late}j de retard — fin prévue {p.date_fin_prevue.strftime("%d/%m/%Y")}',
            'url': f'/projets/{p.pk}/',
            'time': f'{days_late}j',
        })

    # Low disbursement projects
    proj_ids = list(projets_qs.filter(statut='en_cours').values_list('id', flat=True))
    dec_by_project = {
        row['financement__projet_id']: float(row['total'] or 0)
        for row in Decaissement.objects.filter(
            financement__projet_id__in=proj_ids
        ).values('financement__projet_id').annotate(total=Sum('montant'))
    }
    for p in projets_qs.filter(statut='en_cours').only('id', 'code', 'montant_total'):
        if len(notifs) >= 10:
            break
        total_dec = dec_by_project.get(p.id, 0)
        mt = float(p.montant_total or 0)
        if mt > 0:
            taux = round(total_dec / mt * 100, 1)
            if taux < 20:
                notifs.append({
                    'type': 'alert',
                    'icon': 'trending_down',
                    'title': f'{p.code} — décaissement faible',
                    'message': f'Taux de décaissement: {taux}%',
                    'url': f'/projets/{p.pk}/',
                    'time': '',
                })

    # Recent modifications from ActivityLog
    from accounts.models import ActivityLog
    recent_logs = ActivityLog.objects.filter(
        action__in=['create', 'update'],
        timestamp__gte=timezone.now() - timezone.timedelta(days=7)
    ).select_related('user').order_by('-timestamp')[:3]
    for log in recent_logs:
        notifs.append({
            'type': 'info',
            'icon': 'edit_note',
            'title': f'{log.get_action_display()} — {log.model_name}',
            'message': f'{log.object_repr[:50]} par {log.user.get_full_name() or log.user.username if log.user else "?"}',
            'url': '#',
            'time': f'{(today - log.timestamp.date()).days}j' if (today - log.timestamp.date()).days > 0 else "Aujourd'hui",
        })

    return JsonResponse({'notifications': notifs[:10], 'count': len(notifs)})


# ============================================================
# Tableau de bord Ministre (vue ultra-synthétique)
# ============================================================

@login_required_custom
def ministre_dashboard(request):
    """Vue synthétique de décision pour le Ministre (CDC §4.2)."""
    today = timezone.now().date()

    # Tous les projets/bailleurs/financements (Ministre voit tout)
    projets_qs = Projet.objects.all()
    bailleurs_qs = Bailleur.objects.all()
    financements_qs = Financement.objects.all()
    decaissements_qs = Decaissement.objects.all()

    # ── KPIs synthétiques (4 à 8 max) ──
    total_projets = projets_qs.count()
    projets_actifs = projets_qs.filter(statut='en_cours').count()
    projets_en_retard = projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today).count()
    total_bailleurs = bailleurs_qs.count()

    total_engage = float(financements_qs.aggregate(t=Sum('montant_engage'))['t'] or 0)
    total_decaisse = float(decaissements_qs.aggregate(t=Sum('montant'))['t'] or 0)
    pipeline = total_engage - total_decaisse
    taux_decaissement = round((total_decaisse / total_engage * 100), 1) if total_engage > 0 else 0

    # ── Alertes stratégiques (top 5 critiques) ──
    alertes_retard = list(
        projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today)
        .select_related('bailleur_principal').order_by('date_fin_prevue')[:5]
    )
    for p in alertes_retard:
        p.jours_retard = (today - p.date_fin_prevue).days

    # Décaissement faible
    alertes_decaissement = []
    for p in projets_qs.filter(statut='en_cours').select_related('bailleur_principal')[:50]:
        if p.taux_decaissement < 20 and float(p.total_engage) > 0:
            alertes_decaissement.append(p)
        if len(alertes_decaissement) >= 5:
            break

    # ── Répartition sectorielle (camembert) ──
    repartition_secteur = {}
    for p in projets_qs.select_related('secteur'):
        s = p.secteur.nom if p.secteur else 'Non défini'
        repartition_secteur[s] = repartition_secteur.get(s, 0) + 1

    # ── Top 5 bailleurs par engagement ──
    top_bailleurs = []
    for b in bailleurs_qs:
        engage = float(financements_qs.filter(bailleur_id=b.id).aggregate(t=Sum('montant_engage'))['t'] or 0)
        if engage > 0:
            decaisse = float(decaissements_qs.filter(financement__bailleur_id=b.id).aggregate(t=Sum('montant'))['t'] or 0)
            top_bailleurs.append({
                'sigle': b.sigle or b.nom[:20],
                'nom': b.nom,
                'engage': engage,
                'decaisse': decaisse,
                'taux': round((decaisse / engage * 100), 1) if engage > 0 else 0,
            })
    top_bailleurs.sort(key=lambda x: x['engage'], reverse=True)
    top_bailleurs = top_bailleurs[:5]

    # ── Évolution décaissements (12 derniers mois) ──
    from collections import OrderedDict
    evolution = OrderedDict()
    for i in range(11, -1, -1):
        m = today.replace(day=1) - timezone.timedelta(days=i*30)
        key = m.strftime('%Y-%m')
        evolution[key] = 0
    for d in decaissements_qs.filter(date_decaissement__gte=today - timezone.timedelta(days=365)):
        key = d.date_decaissement.strftime('%Y-%m')
        if key in evolution:
            evolution[key] += float(d.montant or 0)

    # ── Carte: nombre de projets par région (PAS de montants) ──
    from collections import defaultdict
    projets_par_region = defaultdict(int)
    for p in projets_qs.filter(statut='en_cours'):
        if p.zone_geographique:
            for z in [s.strip() for s in p.zone_geographique.split(',') if s.strip()]:
                projets_par_region[z] += 1

    context = {
        # KPIs
        'total_projets': total_projets,
        'projets_actifs': projets_actifs,
        'projets_en_retard': projets_en_retard,
        'total_bailleurs': total_bailleurs,
        'total_engage': total_engage,
        'total_decaisse': total_decaisse,
        'pipeline': pipeline,
        'taux_decaissement': taux_decaissement,
        # Alertes
        'alertes_retard': alertes_retard,
        'alertes_decaissement': alertes_decaissement,
        # Graphiques
        'repartition_secteur_json': json.dumps(repartition_secteur),
        'top_bailleurs_json': json.dumps(top_bailleurs),
        'top_bailleurs': top_bailleurs,
        'evolution_json': json.dumps(list(evolution.items())),
        'projets_par_region_json': json.dumps(dict(projets_par_region)),
    }
    return render(request, 'dashboard/ministre.html', context)


# ============================================================
# Tableau de bord ponctualité (CDC §5.10, demande ENSEA)
# ============================================================

@login_required_custom
def ponctualite(request):
    """Suivi de la ponctualité de saisie par structure (Admin/DirCab)."""
    profile = getattr(request.user, 'profile', None)
    if not (request.user.is_superuser or (profile and profile.is_directeur)):
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, "Accès réservé à l'administrateur et au DirCab.")
        return redirect('dashboard:index')

    today = timezone.now().date()
    seuil_jours = int(request.GET.get('seuil', 60))

    # Regroupement par structure responsable
    from collections import defaultdict
    par_structure = defaultdict(lambda: {'total': 0, 'a_jour': 0, 'en_retard': 0, 'projets': []})
    for p in Projet.objects.exclude(structure_responsable='').select_related('bailleur_principal'):
        key = p.structure_responsable or 'Non renseigné'
        jours = p.jours_depuis_modification or 0
        par_structure[key]['total'] += 1
        if jours <= seuil_jours:
            par_structure[key]['a_jour'] += 1
        else:
            par_structure[key]['en_retard'] += 1
        par_structure[key]['projets'].append({
            'code': p.code, 'titre': p.titre, 'jours': jours,
            'pk': p.pk, 'a_jour': jours <= seuil_jours,
        })

    structures = []
    for nom, data in par_structure.items():
        taux = round(data['a_jour'] / data['total'] * 100, 1) if data['total'] > 0 else 0
        structures.append({
            'nom': nom,
            'total': data['total'],
            'a_jour': data['a_jour'],
            'en_retard': data['en_retard'],
            'taux_ponctualite': taux,
            'projets': data['projets'],
        })
    structures.sort(key=lambda s: s['taux_ponctualite'])

    # Stats globales
    nb_total = sum(s['total'] for s in structures)
    nb_a_jour = sum(s['a_jour'] for s in structures)
    taux_global = round(nb_a_jour / nb_total * 100, 1) if nb_total > 0 else 0

    return render(request, 'dashboard/ponctualite.html', {
        'structures': structures,
        'seuil_jours': seuil_jours,
        'nb_total': nb_total,
        'nb_a_jour': nb_a_jour,
        'taux_global': taux_global,
    })


# ============================================================
# Export synthèse KPI (CDC §5.7, demande ENSEA)
# ============================================================

@login_required_custom
def exporter_kpi(request):
    """Exporte la synthèse KPI du dashboard en Excel (CDC §5.7)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = timezone.now().date()
    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets_qs = Projet.objects.all()
        bailleurs_qs = Bailleur.objects.all()
        financements_qs = Financement.objects.all()
        decaissements_qs = Decaissement.objects.all()
    else:
        bailleurs_qs = Bailleur.objects.filter(pk__in=bailleur_ids)
        projets_qs = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()
        financements_qs = Financement.objects.filter(bailleur_id__in=bailleur_ids)
        decaissements_qs = Decaissement.objects.filter(financement__bailleur_id__in=bailleur_ids)

    total_projets = projets_qs.count()
    projets_actifs = projets_qs.filter(statut='en_cours').count()
    projets_retard = projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today).count()
    total_engage = float(financements_qs.aggregate(t=Sum('montant_engage'))['t'] or 0)
    total_decaisse = float(decaissements_qs.aggregate(t=Sum('montant'))['t'] or 0)
    pipeline = total_engage - total_decaisse
    taux = round(total_decaisse / total_engage * 100, 1) if total_engage > 0 else 0
    nb_bailleurs = bailleurs_qs.count()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Synthèse KPI'

    title_fill = PatternFill('solid', fgColor='F77F00')
    title_font = Font(bold=True, color='FFFFFF', size=14)
    label_font = Font(bold=True, size=10, color='4B5563')
    val_font = Font(bold=True, size=12, color='111827')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:C1')
    c = ws.cell(row=1, column=1, value=f"Synthèse KPI — Plateforme de Suivi des Projets Cofinancés ({today.strftime('%d/%m/%Y')})")
    c.fill = title_fill; c.font = title_font; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    rows = [
        ('Indicateur', 'Valeur', 'Unité'),
        ('Nombre total de projets', total_projets, ''),
        ('Projets actifs (en cours)', projets_actifs, ''),
        ('Projets en retard', projets_retard, ''),
        ('Nombre de bailleurs', nb_bailleurs, ''),
        ('Montant total engagé', total_engage, 'devise origine'),
        ('Montant total décaissé', total_decaisse, 'devise origine'),
        ('Pipeline (engagé non décaissé)', pipeline, 'devise origine'),
        ('Taux de décaissement global', taux, '%'),
    ]
    header_fill = PatternFill('solid', fgColor='FED7AA')
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border
            if i == 2:
                cell.fill = header_fill
                cell.font = label_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = val_font if j == 2 else label_font
                cell.alignment = Alignment(vertical='center', horizontal='left' if j != 2 else 'right')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'synthese_kpi_{today.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ============================================================
# Export PDF Synthèse KPI (reportlab)
# ============================================================

@login_required_custom
def exporter_kpi_pdf(request):
    """Exporte la synthèse KPI en PDF style rapport officiel."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io

    today = timezone.now().date()
    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets_qs = Projet.objects.all()
        bailleurs_qs = Bailleur.objects.all()
        financements_qs = Financement.objects.all()
        decaissements_qs = Decaissement.objects.all()
    else:
        bailleurs_qs = Bailleur.objects.filter(pk__in=bailleur_ids)
        projets_qs = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()
        financements_qs = Financement.objects.filter(bailleur_id__in=bailleur_ids)
        decaissements_qs = Decaissement.objects.filter(financement__bailleur_id__in=bailleur_ids)

    total_projets = projets_qs.count()
    projets_actifs = projets_qs.filter(statut='en_cours').count()
    projets_retard = projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today).count()
    total_engage = float(financements_qs.aggregate(t=Sum('montant_engage'))['t'] or 0)
    total_decaisse = float(decaissements_qs.aggregate(t=Sum('montant'))['t'] or 0)
    pipeline = total_engage - total_decaisse
    taux = round(total_decaisse / total_engage * 100, 1) if total_engage > 0 else 0
    nb_bailleurs = bailleurs_qs.count()

    # Projets en retard
    projets_retard_list = list(
        projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today)
        .select_related('bailleur_principal', 'secteur')
        .order_by('date_fin_prevue')[:20]
    )

    CI_ORANGE = colors.HexColor('#F77F00')
    CI_GREEN = colors.HexColor('#009A44')
    CI_DARK = colors.HexColor('#1E293B')
    CI_LIGHT = colors.HexColor('#F8FAFC')
    CI_BORDER = colors.HexColor('#E2E8F0')
    WHITE = colors.white

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2.5*cm, bottomMargin=2.5*cm,
        title=f'Synthèse KPI — {today.strftime("%d/%m/%Y")}',
        author='Plateforme Suivi Projets Bailleurs — RCI',
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=18, textColor=WHITE,
                              alignment=TA_CENTER, leading=22, spaceAfter=0)
    s_subtitle = ParagraphStyle('Sub', fontName='Helvetica', fontSize=10, textColor=CI_DARK,
                                 alignment=TA_CENTER, spaceAfter=12)
    s_section = ParagraphStyle('Sec', fontName='Helvetica-Bold', fontSize=11, textColor=CI_ORANGE,
                                spaceAfter=6, spaceBefore=14, leading=14)
    s_body = ParagraphStyle('Body', fontName='Helvetica', fontSize=9, textColor=CI_DARK,
                             spaceAfter=4, leading=13)
    s_footer = ParagraphStyle('Footer', fontName='Helvetica', fontSize=7.5,
                               textColor=colors.HexColor('#94A3B8'), alignment=TA_CENTER)

    elems = []

    # ── Bannière d'en-tête ──
    header_data = [[Paragraph(
        'RÉPUBLIQUE DE CÔTE D\'IVOIRE — Ministère du Plan et du Développement<br/>'
        '<font size="14">RAPPORT DE SYNTHÈSE KPI</font><br/>'
        f'<font size="9">Plateforme de suivi des projets cofinancés par les bailleurs de fonds</font>',
        s_title
    )]]
    header_table = Table(header_data, colWidths=[17*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), CI_ORANGE),
        ('ROWPADDING', (0, 0), (-1, -1), 10),
        ('ROUNDEDCORNERS', [6]),
    ]))
    elems.append(header_table)
    elems.append(Spacer(1, 0.3*cm))

    # Bandeau vert + blanc + orange (drapeau CI)
    flag_data = [['']]
    flag_table = Table(flag_data, colWidths=[17*cm], rowHeights=[4])
    flag_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), CI_GREEN),
    ]))
    elems.append(flag_table)
    elems.append(Spacer(1, 0.4*cm))

    elems.append(Paragraph(
        f'Date d\'édition : <b>{today.strftime("%d %B %Y")}</b> &nbsp;|&nbsp; '
        f'Généré par : <b>{request.user.get_full_name() or request.user.username}</b>',
        s_subtitle
    ))
    elems.append(HRFlowable(width='100%', thickness=1, color=CI_BORDER, spaceAfter=10))

    # ── KPIs principaux ──
    elems.append(Paragraph('I. INDICATEURS CLÉS DE PERFORMANCE', s_section))

    kpi_data = [
        ['Indicateur', 'Valeur', 'Observation'],
        ['Nombre total de projets', str(total_projets), ''],
        ['Projets actifs (en cours d\'exécution)', str(projets_actifs), f'{round(projets_actifs/total_projets*100, 1) if total_projets else 0} % du portefeuille'],
        ['Projets en situation de retard', str(projets_retard), 'À surveiller'],
        ['Nombre de bailleurs actifs', str(nb_bailleurs), ''],
        ['Montant total engagé', f'{total_engage:,.0f}', 'En devise d\'origine'],
        ['Montant total décaissé', f'{total_decaisse:,.0f}', 'En devise d\'origine'],
        ['Pipeline (engagé non décaissé)', f'{pipeline:,.0f}', 'En devise d\'origine'],
        ['Taux de décaissement global', f'{taux} %', '≥ 70 % = satisfaisant'],
    ]
    kpi_table = Table(kpi_data, colWidths=[8*cm, 4*cm, 5*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), CI_ORANGE),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [CI_LIGHT, WHITE]),
        ('GRID', (0, 0), (-1, -1), 0.5, CI_BORDER),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ('ROWPADDING', (0, 0), (-1, -1), 5),
        # Ligne retard en rouge si > 0
    ]))
    if projets_retard > 0:
        retard_idx = kpi_data.index(['Projets en situation de retard', str(projets_retard), 'À surveiller'])
        kpi_table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, retard_idx), (-1, retard_idx), colors.HexColor('#DC2626')),
            ('FONTNAME', (0, retard_idx), (-1, retard_idx), 'Helvetica-Bold'),
        ]))
    elems.append(kpi_table)
    elems.append(Spacer(1, 0.5*cm))

    # ── Projets en retard ──
    if projets_retard_list:
        elems.append(Paragraph('II. PROJETS EN RETARD D\'EXÉCUTION', s_section))
        elems.append(Paragraph(
            f'Les {len(projets_retard_list)} projets ci-dessous ont dépassé leur date de fin prévisionnelle '
            'et nécessitent une attention particulière.', s_body
        ))
        elems.append(Spacer(1, 0.2*cm))

        retard_data = [['Code', 'Titre du projet', 'Bailleur', 'Fin prévue', 'Retard (j.)']]
        for p in projets_retard_list:
            jours = (today - p.date_fin_prevue).days if p.date_fin_prevue else 0
            retard_data.append([
                p.code,
                (p.titre[:40] + '…') if len(p.titre) > 40 else p.titre,
                (p.bailleur_principal.sigle or p.bailleur_principal.nom[:15]) if p.bailleur_principal else '-',
                p.date_fin_prevue.strftime('%d/%m/%Y') if p.date_fin_prevue else '-',
                str(jours),
            ])
        retard_table = Table(retard_data, colWidths=[2.5*cm, 6.5*cm, 2.5*cm, 2.5*cm, 2*cm])
        retard_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), CI_DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFF7F7'), WHITE]),
            ('GRID', (0, 0), (-1, -1), 0.5, CI_BORDER),
            ('TEXTCOLOR', (4, 1), (4, -1), colors.HexColor('#DC2626')),
            ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),
            ('ROWPADDING', (0, 0), (-1, -1), 4),
        ]))
        elems.append(retard_table)
        elems.append(Spacer(1, 0.5*cm))

    # ── Pied de page ──
    elems.append(HRFlowable(width='100%', thickness=0.5, color=CI_BORDER, spaceBefore=10))
    elems.append(Paragraph(
        'Document généré automatiquement par la Plateforme de Suivi des Projets Cofinancés — '
        'Ministère du Plan et du Développement — République de Côte d\'Ivoire',
        s_footer
    ))

    doc.build(elems)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    filename = f'rapport_kpi_{today.strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# Export PDF Rapport retards (reportlab)
# ============================================================

@login_required_custom
def exporter_rapport_retards_pdf(request):
    """Rapport PDF des projets en retard — style rapport officiel."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io

    today = timezone.now().date()
    bailleur_ids = _get_user_bailleur_ids(request.user)
    projets_qs = Projet.objects.all() if bailleur_ids is None else Projet.objects.filter(
        Q(bailleur_principal_id__in=bailleur_ids) |
        Q(financements__bailleur_id__in=bailleur_ids)
    ).distinct()

    retards = list(
        projets_qs.filter(statut='en_cours', date_fin_prevue__lt=today)
        .select_related('bailleur_principal', 'secteur')
        .order_by('date_fin_prevue')
    )
    for p in retards:
        p.jours_retard = (today - p.date_fin_prevue).days if p.date_fin_prevue else 0

    CI_ORANGE = colors.HexColor('#F77F00')
    CI_GREEN = colors.HexColor('#009A44')
    CI_DARK = colors.HexColor('#1E293B')
    CI_LIGHT = colors.HexColor('#F8FAFC')
    CI_BORDER = colors.HexColor('#E2E8F0')
    WHITE = colors.white
    RED = colors.HexColor('#DC2626')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2.5*cm, bottomMargin=2.5*cm)

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('T', fontName='Helvetica-Bold', fontSize=16, textColor=WHITE,
                              alignment=TA_CENTER, leading=22, spaceAfter=0)
    s_section = ParagraphStyle('S', fontName='Helvetica-Bold', fontSize=11, textColor=CI_ORANGE,
                                spaceAfter=6, spaceBefore=14)
    s_body = ParagraphStyle('B', fontName='Helvetica', fontSize=9, textColor=CI_DARK,
                             spaceAfter=4, leading=13)
    s_footer = ParagraphStyle('F', fontName='Helvetica', fontSize=7.5,
                               textColor=colors.HexColor('#94A3B8'), alignment=TA_CENTER)

    elems = []

    # ── En-tête ──
    header_data = [[Paragraph(
        'RÉPUBLIQUE DE CÔTE D\'IVOIRE — Ministère du Plan et du Développement<br/>'
        '<font size="14">RAPPORT SUR LES PROJETS EN RETARD</font><br/>'
        f'<font size="9">Édité le {today.strftime("%d/%m/%Y")}</font>',
        s_title
    )]]
    ht = Table(header_data, colWidths=[17*cm])
    ht.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), CI_ORANGE),
        ('ROWPADDING', (0, 0), (-1, -1), 10),
    ]))
    elems.append(ht)
    elems.append(Spacer(1, 0.15*cm))
    flag_table = Table([['']], colWidths=[17*cm], rowHeights=[4])
    flag_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), CI_GREEN)]))
    elems.append(flag_table)
    elems.append(Spacer(1, 0.4*cm))

    elems.append(Paragraph(
        f'Ce rapport recense <b>{len(retards)} projet(s)</b> en situation de retard à la date du '
        f'<b>{today.strftime("%d/%m/%Y")}</b>.', s_body))
    elems.append(HRFlowable(width='100%', thickness=1, color=CI_BORDER, spaceBefore=6, spaceAfter=10))

    if retards:
        data = [['Code', 'Titre du projet', 'Bailleur', 'Secteur', 'Fin prévue', 'Retard (j.)', 'Motif']]
        for p in retards:
            motif = p.motif_retard_categorie or '-'
            data.append([
                p.code,
                (p.titre[:35] + '…') if len(p.titre) > 35 else p.titre,
                (p.bailleur_principal.sigle or p.bailleur_principal.nom[:12]) if p.bailleur_principal else '-',
                (p.secteur.nom[:12] if p.secteur else '-'),
                p.date_fin_prevue.strftime('%d/%m/%Y') if p.date_fin_prevue else '-',
                str(p.jours_retard),
                motif.capitalize(),
            ])
        t = Table(data, colWidths=[2*cm, 5.5*cm, 2*cm, 2*cm, 2*cm, 1.5*cm, 2*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), CI_DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFF7F7'), WHITE]),
            ('GRID', (0, 0), (-1, -1), 0.4, CI_BORDER),
            ('TEXTCOLOR', (5, 1), (5, -1), RED),
            ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
            ('ALIGN', (5, 0), (5, -1), 'CENTER'),
            ('ROWPADDING', (0, 0), (-1, -1), 4),
        ]))
        elems.append(t)
    else:
        elems.append(Paragraph('Aucun projet en retard à ce jour.', s_body))

    elems.append(Spacer(1, 0.5*cm))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=CI_BORDER))
    elems.append(Paragraph(
        'Document généré automatiquement par la Plateforme de Suivi des Projets Cofinancés — '
        'Ministère du Plan et du Développement — République de Côte d\'Ivoire', s_footer))

    doc.build(elems)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    filename = f'rapport_retards_{today.strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
