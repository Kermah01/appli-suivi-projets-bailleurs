import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg
from django.http import HttpResponse
from django.utils import timezone
from .models import Projet, Secteur, Programme, PieceJointe, ResponsableLocal, CommentaireProjet
from .forms import ProjetForm, ProgrammeForm, PieceJointeForm, ResponsableLocalForm
from bailleurs.models import Bailleur
from financements.models import Financement, Decaissement
from accounts.decorators import login_required_custom, edit_permission_required
from accounts.models import ActivityLog


@login_required_custom
def liste(request):
    query = request.GET.get('q', '')
    statut_filter = request.GET.get('statut', '')
    secteur_filter = request.GET.get('secteur', '')
    bailleur_filter = request.GET.get('bailleur', '')
    en_retard = request.GET.get('en_retard', '')
    from dashboard.views import _get_user_bailleur_ids
    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets = Projet.objects.all()
    else:
        projets = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()
    projets = projets.select_related('secteur', 'bailleur_principal').prefetch_related('financements__bailleur')

    if query:
        projets = projets.filter(Q(titre__icontains=query) | Q(code__icontains=query))
    if statut_filter:
        projets = projets.filter(statut=statut_filter)
    if en_retard:
        projets = projets.filter(statut='en_cours', date_fin_prevue__lt=timezone.now().date())
    if secteur_filter:
        projets = projets.filter(secteur_id=secteur_filter)
    if bailleur_filter:
        projets = projets.filter(
            Q(bailleur_principal_id=bailleur_filter) |
            Q(financements__bailleur_id=bailleur_filter)
        ).distinct()

    from bailleurs.models import Bailleur
    context = {
        'projets': projets,
        'query': query,
        'statut_filter': statut_filter,
        'secteur_filter': secteur_filter,
        'bailleur_filter': bailleur_filter,
        'en_retard': en_retard,
        'statuts': Projet.STATUT_CHOICES,
        'secteurs': Secteur.objects.all(),
        'bailleurs': Bailleur.objects.all(),
    }
    return render(request, 'projets/liste.html', context)


@login_required_custom
def detail(request, pk):
    projet = get_object_or_404(
        Projet.objects.select_related('secteur', 'bailleur_principal', 'programme').prefetch_related(
            'objectifs_pnd__pilier', 'financements__bailleur', 'financements__decaissements',
            'pieces_jointes', 'responsables_locaux', 'commentaires__auteur'
        ),
        pk=pk
    )
    commentaires = projet.commentaires.all()
    context = {
        'projet': projet,
        'piece_form': PieceJointeForm(),
        'responsable_form': ResponsableLocalForm(),
        'commentaires': commentaires,
        'type_commentaire_choices': CommentaireProjet.TYPE_CHOICES,
        'niveau_commentaire_choices': CommentaireProjet.NIVEAU_CHOICES,
    }
    return render(request, 'projets/detail.html', context)


@login_required_custom
def ajouter_commentaire(request, pk):
    projet = get_object_or_404(Projet, pk=pk)
    if request.method == 'POST':
        contenu = request.POST.get('contenu', '').strip()
        type_c = request.POST.get('type_commentaire', 'observation')
        niveau = request.POST.get('niveau', 'information')
        if contenu:
            from django.utils import timezone as tz
            CommentaireProjet.objects.create(
                projet=projet,
                type_commentaire=type_c,
                niveau=niveau,
                contenu=contenu,
                date_commentaire=tz.now().date(),
                source='manuel',
                auteur=request.user,
            )
            messages.success(request, 'Commentaire ajouté.')
        else:
            messages.error(request, 'Le contenu ne peut pas être vide.')
    return redirect('projets:detail', pk=pk)


@login_required_custom
def supprimer_commentaire(request, pk, commentaire_pk):
    commentaire = get_object_or_404(CommentaireProjet, pk=commentaire_pk, projet_id=pk)
    if request.user == commentaire.auteur or request.user.is_superuser:
        commentaire.delete()
        messages.success(request, 'Commentaire supprimé.')
    else:
        messages.error(request, 'Permission refusée.')
    return redirect('projets:detail', pk=pk)


def _create_financements_from_json(projet, financements_json_str):
    """Crée les financements à partir du JSON envoyé par le formulaire Alpine.js."""
    if not financements_json_str:
        return 0
    try:
        items = json.loads(financements_json_str)
    except (json.JSONDecodeError, TypeError):
        return 0
    count = 0
    for item in items:
        bailleur_id = item.get('bailleur_id')
        montant = item.get('montant')
        if not bailleur_id or not montant:
            continue
        try:
            bailleur = Bailleur.objects.get(pk=bailleur_id)
        except Bailleur.DoesNotExist:
            continue
        type_fin = item.get('type_financement', 'don') or 'don'
        devise = item.get('devise', projet.devise) or projet.devise
        Financement.objects.update_or_create(
            projet=projet, bailleur=bailleur, type_financement=type_fin,
            defaults={
                'montant_engage': montant,
                'devise': devise,
                'date_accord': projet.date_signature,
            }
        )
        count += 1
    return count


@edit_permission_required
def creer(request):
    bailleur_id = request.GET.get('bailleur')
    if request.method == 'POST':
        form = ProjetForm(request.POST)
        if form.is_valid():
            projet = form.save()
            _create_financements_from_json(projet, form.cleaned_data.get('financements_json'))
            ActivityLog.log(request.user, 'create', 'Projet', f'[{projet.code}] {projet.titre}', object_id=projet.pk)
            messages.success(request, f'Projet "{projet.titre}" créé avec succès.')
            return redirect('projets:detail', pk=projet.pk)
    else:
        initial = {}
        if bailleur_id:
            initial['bailleur_principal'] = bailleur_id
        form = ProjetForm(initial=initial)
    return render(request, 'projets/form.html', {
        'form': form,
        'titre': 'Nouveau projet',
        'is_creation': True,
    })


@edit_permission_required
def modifier(request, pk):
    projet = get_object_or_404(Projet, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce bailleur.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        form = ProjetForm(request.POST, instance=projet)
        if form.is_valid():
            form.save()
            fin_json = form.cleaned_data.get('financements_json')
            if fin_json:
                _create_financements_from_json(projet, fin_json)
            ActivityLog.log(request.user, 'update', 'Projet', f'[{projet.code}] {projet.titre}', object_id=projet.pk)
            messages.success(request, f'Projet "{projet.titre}" modifié avec succès.')
            return redirect('projets:detail', pk=projet.pk)
    else:
        form = ProjetForm(instance=projet)
    return render(request, 'projets/form.html', {
        'form': form, 'titre': 'Modifier le projet',
        'projet': projet, 'is_creation': False,
    })


@login_required_custom
def exporter_excel(request):
    """Exporte la liste des projets filtrée en Excel via openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from dashboard.views import _get_user_bailleur_ids
    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets = Projet.objects.all()
    else:
        projets = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()

    query = request.GET.get('q', '')
    statut_filter = request.GET.get('statut', '')
    secteur_filter = request.GET.get('secteur', '')
    bailleur_filter = request.GET.get('bailleur', '')
    if query:
        projets = projets.filter(Q(titre__icontains=query) | Q(code__icontains=query))
    if statut_filter:
        projets = projets.filter(statut=statut_filter)
    if secteur_filter:
        projets = projets.filter(secteur_id=secteur_filter)
    if bailleur_filter:
        projets = projets.filter(
            Q(bailleur_principal_id=bailleur_filter) |
            Q(financements__bailleur_id=bailleur_filter)
        ).distinct()

    projets = projets.select_related('secteur', 'bailleur_principal')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projets'

    header_fill = PatternFill('solid', fgColor='F77F00')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Code', 'Titre du projet', 'Bailleur principal', 'Secteur',
        'Statut', 'Montant total', 'Devise', 'Taux avancement (%)',
        'Zone géographique', 'Responsable', 'Structure responsable',
        'Date signature', 'Date début', 'Date fin prévue', 'Date fin effective',
        'Montant décaissé', 'Taux décaissement (%)', 'Motif retard',
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_idx, p in enumerate(projets, 2):
        row = [
            p.code,
            p.titre,
            (p.bailleur_principal.sigle or p.bailleur_principal.nom) if p.bailleur_principal else '',
            p.secteur.nom if p.secteur else '',
            p.get_statut_display(),
            float(p.montant_total),
            p.devise,
            float(p.taux_avancement),
            p.zone_geographique,
            p.responsable,
            p.structure_responsable,
            p.date_signature.strftime('%d/%m/%Y') if p.date_signature else '',
            p.date_debut.strftime('%d/%m/%Y') if p.date_debut else '',
            p.date_fin_prevue.strftime('%d/%m/%Y') if p.date_fin_prevue else '',
            p.date_fin_effective.strftime('%d/%m/%Y') if p.date_fin_effective else '',
            float(p.total_decaisse),
            float(p.taux_decaissement),
            p.motif_retard,
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FFF7ED')

    col_widths = [12, 45, 18, 18, 16, 15, 8, 12, 22, 22, 22, 14, 14, 14, 14, 15, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'projets_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@edit_permission_required
def supprimer_lot(request):
    """Suppression en masse de plusieurs projets sélectionnés."""
    if request.method != 'POST':
        return redirect('projets:liste')

    ids = request.POST.getlist('ids[]')
    if not ids:
        messages.warning(request, "Aucun projet sélectionné.")
        return redirect('projets:liste')

    profile = getattr(request.user, 'profile', None)
    supprimés, refusés = 0, 0
    for projet in Projet.objects.filter(pk__in=ids):
        if profile and not profile.can_edit_projet(projet):
            refusés += 1
            continue
        titre, code = projet.titre, projet.code
        projet.delete()
        ActivityLog.log(request.user, 'delete', 'Projet', f'[{code}] {titre}')
        supprimés += 1

    if supprimés:
        messages.success(request, f"{supprimés} projet{'s' if supprimés > 1 else ''} supprimé{'s' if supprimés > 1 else ''}.")
    if refusés:
        messages.warning(request, f"{refusés} projet{'s' if refusés > 1 else ''} ignoré{'s' if refusés > 1 else ''} (droits insuffisants).")
    return redirect('projets:liste')


@edit_permission_required
def supprimer(request, pk):
    projet = get_object_or_404(Projet, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce bailleur.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        titre = projet.titre
        code = projet.code
        projet.delete()
        ActivityLog.log(request.user, 'delete', 'Projet', f'[{code}] {titre}')
        messages.success(request, f'Projet "{titre}" supprimé.')
        return redirect('projets:liste')
    return render(request, 'projets/confirmer_suppression.html', {'projet': projet})


# ============================================================
# Programmes — helpers et CRUD
# ============================================================

def _programme_stats(programme, methode='auto'):
    """
    Calcule les métriques effectives d'un programme.
    methode: 'auto' | 'saisi' | 'agrege'
      - auto   : valeur saisie si > 0, sinon agrégat des projets
      - saisi  : toujours la valeur directe du programme
      - agrege : toujours l'agrégat des projets enfants
    """
    agg = programme.projets.aggregate(
        sum_montant=Sum('montant_total'),
        avg_ta=Avg('taux_avancement'),
        avg_taf=Avg('taux_avancement_financier'),
        avg_tdp=Avg('taux_decaissement_prevu_annee'),
    )
    montant_ag = float(agg['sum_montant'] or 0)
    ta_ag  = round(float(agg['avg_ta']  or 0), 1)
    taf_ag = round(float(agg['avg_taf'] or 0), 1)
    tdp_ag = round(float(agg['avg_tdp'] or 0), 1)

    montant_s = float(programme.montant_total or 0)
    ta_s  = float(programme.taux_avancement or 0)
    taf_s = float(programme.taux_avancement_financier or 0)
    tdp_s = float(programme.taux_decaissement_prevu_annee or 0)

    engage   = float(Financement.objects.filter(projet__programme=programme).aggregate(t=Sum('montant_engage'))['t'] or 0)
    decaisse = float(Decaissement.objects.filter(financement__projet__programme=programme).aggregate(t=Sum('montant'))['t'] or 0)
    taux_dec = round(decaisse / engage * 100, 1) if engage > 0 else 0

    def _pick(saisi, agrege):
        if methode == 'saisi':
            return (saisi, 'saisi')
        elif methode == 'agrege':
            return (agrege, 'agrege')
        else:  # auto
            if saisi > 0:
                return (saisi, 'saisi')
            elif agrege > 0:
                return (agrege, 'agrege')
            return (0, 'vide')

    mt,  mt_src  = _pick(montant_s, montant_ag)
    ta,  ta_src  = _pick(ta_s,  ta_ag)
    taf, taf_src = _pick(taf_s, taf_ag)
    tdp, tdp_src = _pick(tdp_s, tdp_ag)

    return {
        'montant_total':           mt,  'montant_total_src':           mt_src,
        'taux_avancement':         ta,  'taux_avancement_src':         ta_src,
        'taux_avancement_financier':taf,'taux_avancement_financier_src':taf_src,
        'taux_decaissement_prevu':  tdp,'taux_decaissement_prevu_src':  tdp_src,
        'montant_engage':  engage,
        'montant_decaisse': decaisse,
        'taux_decaissement': taux_dec,
        'nb_projets': programme.projets.count(),
    }

@login_required_custom
def programme_liste(request):
    query = request.GET.get('q', '')
    programmes = Programme.objects.annotate(
        nb_projets=Count('projets', distinct=True)
    )
    if query:
        programmes = programmes.filter(Q(nom__icontains=query) | Q(code__icontains=query))
    return render(request, 'projets/programme_liste.html', {
        'programmes': programmes, 'query': query,
    })


@login_required_custom
def programme_detail(request, pk):
    from django.utils import timezone as tz
    methode = request.GET.get('methode', 'auto')
    if methode not in ('auto', 'saisi', 'agrege'):
        methode = 'auto'
    programme = get_object_or_404(
        Programme.objects.select_related('secteur', 'bailleur_principal').prefetch_related(
            'projets__bailleur_principal', 'projets__secteur',
            'projets__financements__bailleur', 'projets__financements__decaissements',
        ),
        pk=pk
    )
    stats = _programme_stats(programme, methode)

    # Bailleurs distincts parmi les projets enfants
    from bailleurs.models import Bailleur
    bailleur_ids = programme.projets.values_list('bailleur_principal_id', flat=True).distinct()
    bailleurs_enfants = Bailleur.objects.filter(pk__in=bailleur_ids).exclude(pk__isnull=True)

    # Répartition des financements par bailleur sur les projets du programme
    repartition = []
    total_engage = stats['montant_engage'] or 0
    for b in bailleurs_enfants:
        fin_agg = Financement.objects.filter(
            projet__programme=programme, bailleur=b
        ).aggregate(eng=Sum('montant_engage'))
        dec_agg = Decaissement.objects.filter(
            financement__projet__programme=programme, financement__bailleur=b
        ).aggregate(dec=Sum('montant'))
        eng = float(fin_agg['eng'] or 0)
        dec = float(dec_agg['dec'] or 0)
        part = round(eng / total_engage * 100, 1) if total_engage > 0 else 0
        taux_d = round(dec / eng * 100, 1) if eng > 0 else 0
        repartition.append({
            'bailleur': b,
            'sigle': b.sigle or b.nom,
            'montant_engage': eng,
            'montant_decaisse': dec,
            'part_pct': part,
            'taux_decaissement': taux_d,
            'devise': 'FCFA',
        })

    # Est en retard ?
    aujourd_hui = tz.now().date()
    est_en_retard = (
        programme.date_fin and
        programme.date_fin < aujourd_hui and
        programme.statut == 'en_cours'
    )
    # Projets en retard
    projets_en_retard = programme.projets.filter(
        statut='en_cours', date_fin_prevue__lt=aujourd_hui
    )

    # Ecart objectif annuel
    ecart_taux_annuel = None
    if stats['taux_decaissement_prevu'] and stats['taux_decaissement_prevu'] > 0:
        ecart_taux_annuel = round(stats['taux_decaissement'] - stats['taux_decaissement_prevu'], 1)

    reste_a_decaisser = max(0, stats['montant_engage'] - stats['montant_decaisse'])

    return render(request, 'projets/programme_detail.html', {
        'programme': programme,
        'stats': stats,
        'methode': methode,
        'repartition': repartition,
        'est_en_retard': est_en_retard,
        'projets_en_retard': projets_en_retard,
        'ecart_taux_annuel': ecart_taux_annuel,
        'reste_a_decaisser': reste_a_decaisser,
        'est_cofinance': len(repartition) > 1,
    })


@edit_permission_required
def programme_creer(request):
    if request.method == 'POST':
        form = ProgrammeForm(request.POST)
        if form.is_valid():
            programme = form.save()
            ActivityLog.log(request.user, 'create', 'Programme', f'[{programme.code}] {programme.nom}', object_id=programme.pk)
            messages.success(request, f'Programme "{programme.nom}" créé.')
            return redirect('projets:programme_detail', pk=programme.pk)
    else:
        form = ProgrammeForm()
    return render(request, 'projets/programme_form.html', {
        'form': form, 'titre': 'Nouveau programme', 'is_creation': True,
    })


@edit_permission_required
def programme_modifier(request, pk):
    programme = get_object_or_404(Programme, pk=pk)
    if request.method == 'POST':
        form = ProgrammeForm(request.POST, instance=programme)
        if form.is_valid():
            form.save()
            ActivityLog.log(request.user, 'update', 'Programme', f'[{programme.code}] {programme.nom}', object_id=programme.pk)
            messages.success(request, f'Programme "{programme.nom}" modifié.')
            return redirect('projets:programme_detail', pk=programme.pk)
    else:
        form = ProgrammeForm(instance=programme)
    return render(request, 'projets/programme_form.html', {
        'form': form, 'titre': f'Modifier {programme.nom}',
        'programme': programme, 'is_creation': False,
    })


@edit_permission_required
def programme_supprimer(request, pk):
    programme = get_object_or_404(Programme, pk=pk)
    if request.method == 'POST':
        nom = programme.nom
        programme.delete()
        ActivityLog.log(request.user, 'delete', 'Programme', nom)
        messages.success(request, f'Programme "{nom}" supprimé.')
        return redirect('projets:programme_liste')
    return render(request, 'projets/programme_confirmer_suppression.html', {'programme': programme})


# ============================================================
# Pièces jointes
# ============================================================

@edit_permission_required
def piece_jointe_ajouter(request, projet_pk):
    projet = get_object_or_404(Projet, pk=projet_pk)
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce projet.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        form = PieceJointeForm(request.POST, request.FILES)
        if form.is_valid():
            piece = form.save(commit=False)
            piece.projet = projet
            piece.uploaded_by = request.user
            piece.save()
            ActivityLog.log(request.user, 'create', 'PieceJointe', f'{piece.titre} → {projet.code}', object_id=piece.pk)
            messages.success(request, f'Pièce jointe "{piece.titre}" ajoutée.')
        else:
            messages.error(request, "Erreur lors de l'ajout : " + str(form.errors))
    return redirect('projets:detail', pk=projet.pk)


@edit_permission_required
def piece_jointe_supprimer(request, pk):
    piece = get_object_or_404(PieceJointe, pk=pk)
    projet = piece.projet
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce projet.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        titre = piece.titre
        if piece.fichier:
            try:
                piece.fichier.delete(save=False)
            except Exception:
                pass
        piece.delete()
        ActivityLog.log(request.user, 'delete', 'PieceJointe', titre)
        messages.success(request, f'Pièce jointe "{titre}" supprimée.')
    return redirect('projets:detail', pk=projet.pk)


# ============================================================
# Responsables locaux
# ============================================================

@edit_permission_required
def responsable_ajouter(request, projet_pk):
    projet = get_object_or_404(Projet, pk=projet_pk)
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce projet.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        form = ResponsableLocalForm(request.POST)
        if form.is_valid():
            resp = form.save(commit=False)
            resp.projet = projet
            resp.save()
            ActivityLog.log(request.user, 'create', 'ResponsableLocal', f'{resp.nom} → {projet.code}', object_id=resp.pk)
            messages.success(request, f'Responsable local "{resp.nom}" ajouté.')
        else:
            messages.error(request, "Erreur : " + str(form.errors))
    return redirect('projets:detail', pk=projet.pk)


@edit_permission_required
def responsable_supprimer(request, pk):
    resp = get_object_or_404(ResponsableLocal, pk=pk)
    projet = resp.projet
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.can_edit_projet(projet):
        messages.error(request, "Vous n'êtes pas point focal de ce projet.")
        return redirect('projets:detail', pk=projet.pk)
    if request.method == 'POST':
        nom = resp.nom
        resp.delete()
        ActivityLog.log(request.user, 'delete', 'ResponsableLocal', nom)
        messages.success(request, f'Responsable "{nom}" supprimé.')
    return redirect('projets:detail', pk=projet.pk)


# ============================================================
# Export rapport retards
# ============================================================

@login_required_custom
def exporter_rapport_retards(request):
    """Exporte le rapport des projets en retard avec motifs (CDC §5.7)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from dashboard.views import _get_user_bailleur_ids
    bailleur_ids = _get_user_bailleur_ids(request.user)
    if bailleur_ids is None:
        projets = Projet.objects.all()
    else:
        projets = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct()
    today = timezone.now().date()
    projets = projets.filter(statut='en_cours', date_fin_prevue__lt=today).select_related('secteur', 'bailleur_principal')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rapport retards'

    header_fill = PatternFill('solid', fgColor='DC2626')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Code', 'Titre', 'Bailleur principal', 'Secteur',
        'Date fin prévue', 'Jours de retard',
        'Catégorie motif', 'Détail motif',
        'Taux décaissement (%)', 'Responsable', 'Structure',
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

    for row_idx, p in enumerate(projets, 2):
        jours_retard = (today - p.date_fin_prevue).days if p.date_fin_prevue else 0
        row = [
            p.code, p.titre,
            (p.bailleur_principal.sigle or p.bailleur_principal.nom) if p.bailleur_principal else '',
            p.secteur.nom if p.secteur else '',
            p.date_fin_prevue.strftime('%d/%m/%Y') if p.date_fin_prevue else '',
            jours_retard,
            p.get_motif_retard_categorie_display() if p.motif_retard_categorie else 'Non renseigné',
            p.motif_retard or '',
            float(p.taux_decaissement),
            p.responsable, p.structure_responsable,
        ]
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FEF2F2')

    widths = [12, 45, 18, 18, 14, 12, 18, 35, 14, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'rapport_retards_{today.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required_custom
def synthese(request):
    """Vue unifiée Projets + Programmes avec métriques financières et physiques."""
    vue = request.GET.get('vue', 'avancement')
    methode = request.GET.get('methode', 'auto')
    if methode not in ('auto', 'saisi', 'agrege'):
        methode = 'auto'
    from dashboard.views import _get_user_bailleur_ids
    bailleur_ids = _get_user_bailleur_ids(request.user)

    if bailleur_ids is None:
        projets_qs = Projet.objects.select_related('secteur', 'bailleur_principal', 'programme')
    else:
        projets_qs = Projet.objects.filter(
            Q(bailleur_principal_id__in=bailleur_ids) |
            Q(financements__bailleur_id__in=bailleur_ids)
        ).distinct().select_related('secteur', 'bailleur_principal', 'programme')

    projets_qs = projets_qs.annotate(
        montant_engage=Sum('financements__montant_engage'),
        montant_decaisse=Sum('financements__decaissements__montant'),
    )

    programmes_qs = Programme.objects.select_related('secteur', 'bailleur_principal').prefetch_related('projets')

    items_projets = []
    for p in projets_qs.all():
        engage = float(p.montant_engage or 0)
        decaisse = float(p.montant_decaisse or 0)
        taux_dec = round(decaisse / engage * 100, 1) if engage > 0 else 0
        items_projets.append({
            'type': 'projet',
            'code': p.code,
            'nom': p.titre,
            'statut': p.get_statut_display(),
            'secteur': p.secteur.nom if p.secteur else '—',
            'bailleur': p.bailleur_principal.sigle if p.bailleur_principal else '—',
            'montant': float(p.montant_total or 0),
            'devise': p.devise,
            'engage': engage,
            'decaisse': decaisse,
            'taux_avancement': float(p.taux_avancement or 0),
            'taux_avancement_financier': float(p.taux_avancement_financier or 0),
            'taux_decaissement': taux_dec,
            'taux_prevu_annee': float(p.taux_decaissement_prevu_annee or 0),
            'programme': p.programme.nom if p.programme else '—',
            'url': f'/projets/{p.pk}/',
        })

    items_programmes = []
    for prog in programmes_qs.all():
        s = _programme_stats(prog, methode)
        items_programmes.append({
            'type': 'programme',
            'code': prog.code,
            'nom': prog.nom,
            'statut': prog.get_statut_display(),
            'secteur': prog.secteur.nom if prog.secteur else '—',
            'bailleur': prog.bailleur_principal.sigle if prog.bailleur_principal else '—',
            'montant': s['montant_total'],
            'montant_src': s['montant_total_src'],
            'devise': prog.devise,
            'engage': s['montant_engage'],
            'decaisse': s['montant_decaisse'],
            'taux_avancement': s['taux_avancement'],
            'taux_avancement_src': s['taux_avancement_src'],
            'taux_avancement_financier': s['taux_avancement_financier'],
            'taux_avancement_financier_src': s['taux_avancement_financier_src'],
            'taux_decaissement': s['taux_decaissement'],
            'taux_prevu_annee': s['taux_decaissement_prevu'],
            'taux_prevu_annee_src': s['taux_decaissement_prevu_src'],
            'nb_projets': s['nb_projets'],
            'programme': '—',
            'url': f'/projets/programmes/{prog.pk}/',
        })

    context = {
        'vue': vue,
        'methode': methode,
        'items_projets': items_projets,
        'items_programmes': items_programmes,
        'total_projets': len(items_projets),
        'total_programmes': len(items_programmes),
    }
    return render(request, 'projets/synthese.html', context)
