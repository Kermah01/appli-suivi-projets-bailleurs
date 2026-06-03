"""
Script d'extraction du fichier "Portefeuille Groupe de Coordination Arabe"
vers le format du template d'import de l'application.

Structure source (feuille unique 'Portefeuille global') :
  3 sections séparées par des lignes de titre :
    - "OPÉRATIONS EN COURS"               (section_row=20)  → statut en_cours
    - "OPÉRATIONS APPROUVÉES NON DÉMARRÉES" (section_row=59) → statut identification
    - "OPÉRATIONS EN INSTANCE DE CLÔTURE"  (section_row=86)  → statut cloture
  Pour chaque section, lignes d'en-tête sur 2 rangs puis données :
    - Lignes "Secteur …" en col 1 → on met à jour le secteur courant
    - Lignes numériques col 1      → données projet
  Colonnes (1-indexed) :
    1=N°, 2=Nom, 3=Objectifs,
    4=BID, 5=BADEA, 6=FKDEA, 7=OFID, 8=FSD, 9=FADD,
    10=Financements par secteurs (col récap), 11=Part État CI,
    12=Date approbation, 13=Signature accord, 14=Date mise en vigueur,
    15=Date démarrage effectif, 16=Date limite décaissement, 17=Âge,
    18=Taux décaissement %, 19=Taux exécution physique, 20=Observations
"""
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date

SRC = r'C:\Users\MIVS\Downloads\Portefeuille Groupe de Coordination Arabe (1).xlsx'
DST = r'C:\Users\MIVS\Downloads\Import_GCA_Template.xlsx'

# ── Mapping sigle interne du nom ──────────────────────────────────────────
BAILLEUR_SIGLES = {
    'BID': ('BID', 'Banque Islamique de Développement'),
    'BADEA': ('BADEA', 'Banque Arabe pour le Développement Économique en Afrique'),
    'FKDEA': ('FKDEA', 'Fonds Koweïtien pour le Développement Économique Arabe'),
    'OFID': ('OFID', "Fonds OPEP pour le Développement International"),
    'FSD': ('FSD', 'Fonds Saoudien de Développement'),
    'FADD': ('FADD', 'Fonds Abu Dhabi pour le Développement'),
    'FADES': ('FADES', 'Fonds Arabe pour le Développement Économique et Social'),
    'GCA': ('GCA', 'Groupe de Coordination Arabe'),
}

COL_BID   = 4
COL_BADEA = 5
COL_FKDEA = 6
COL_OFID  = 7
COL_FSD   = 8
COL_FADD  = 9

HEADER_ROW = 86  # Ligne d'en-tête dans le fichier source
DATA_START  = 88  # Première ligne de données

# ── Styles ────────────────────────────────────────────────────────────────
DARK  = "1E293B"
ACCENT = "1E3A5F"
GREEN  = "166534"
WHITE  = "FFFFFF"
GRAY   = "F1F5F9"

TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color=WHITE)
HEADER_FONT   = Font(name="Calibri", bold=True, size=10, color=WHITE)
EXAMPLE_FONT  = Font(name="Calibri", size=10, color="475569", italic=True)
THIN_BORDER   = Border(
    left=Side(style='thin', color="CBD5E1"),
    right=Side(style='thin', color="CBD5E1"),
    top=Side(style='thin', color="CBD5E1"),
    bottom=Side(style='thin', color="CBD5E1"),
)

def _d(val):
    """Convertit datetime → date.isoformat ou None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except Exception:
        return None

def _fmt_date(d):
    return d.isoformat() if d else ''

def _num(val):
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    return None

def _pct(val):
    """Retourne un taux en décimal (0-1). Gère 0.77, 77, 'En cours', etc."""
    if val is None:
        return None
    if isinstance(val, str):
        return None
    v = float(val)
    if v > 1.5:  # probablement déjà en %
        return round(v / 100, 4)
    return round(v, 4)

def _extract_sigle_from_name(nom):
    """Extrait le sigle entre parenthèses en fin de nom."""
    if not nom:
        return None
    m = re.search(r'\(([A-Z]{2,6})\)\s*$', str(nom))
    return m.group(1) if m else None

# ── Sections dans le fichier source ─────────────────────────────────────
# (section_header_row, statut_appli, code_prefix)
SECTIONS = [
    (20, 'en_cours',        'EC'),
    (59, 'identification',  'ND'),
    (86, 'cloture',         'CL'),
]

# ── Lecture du fichier source ─────────────────────────────────────────────
print(f"Lecture de {SRC} ...")
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src['Portefeuille global']

# Trouver les limites de chaque section
section_limits = []
for i, (hdr_row, statut, prefix) in enumerate(SECTIONS):
    next_hdr = SECTIONS[i+1][0] if i+1 < len(SECTIONS) else ws_src.max_row + 1
    section_limits.append((hdr_row + 2, next_hdr - 1, statut, prefix))

projets = []
global_n = 0  # numéro séquentiel global

for (data_start, data_end, statut, prefix) in section_limits:
    current_secteur = ''
    for r in range(data_start, data_end + 1):
        v1 = ws_src.cell(row=r, column=1).value
        v2 = ws_src.cell(row=r, column=2).value

        # Lignes secteur : "Secteur …" ou "Secteur" seul
        if v1 is not None and str(v1).strip().lower().startswith('secteur'):
            current_secteur = str(v1).replace('Secteur', '').replace('secteur', '').strip()
            continue

        # Ligne projet : col 1 est un numéro
        if v1 is None or v2 is None:
            continue
        v1_str = str(v1).strip()
        if not v1_str or not (v1_str[0].isdigit()):
            continue
        if v1_str in ('N°',):
            continue

        nom = str(v2).strip()
        sigle = _extract_sigle_from_name(nom) or 'BID'

        mt_bid   = _num(ws_src.cell(row=r, column=4).value)
        mt_badea = _num(ws_src.cell(row=r, column=5).value)
        mt_fkdea = _num(ws_src.cell(row=r, column=6).value)
        mt_ofid  = _num(ws_src.cell(row=r, column=7).value)
        mt_fsd   = _num(ws_src.cell(row=r, column=8).value)
        mt_fadd  = _num(ws_src.cell(row=r, column=9).value)

        bailleur_mts = {
            'BID':   mt_bid,
            'BADEA': mt_badea,
            'FKDEA': mt_fkdea,
            'OFID':  mt_ofid,
            'FSD':   mt_fsd,
            'FADD':  mt_fadd,
        }
        mt_total = sum(v for v in bailleur_mts.values() if v)

        taux_dec  = _pct(ws_src.cell(row=r, column=18).value)
        taux_phys = _pct(ws_src.cell(row=r, column=19).value)
        obs_val = ws_src.cell(row=r, column=20).value

        global_n += 1
        num_str = v1_str.split()[0].rstrip('.')
        code = f"GCA-{prefix}-{global_n:03d}"

        projets.append({
            'num': v1,
            'code': code,
            'nom': nom,
            'sigle': sigle,
            'secteur': current_secteur,
            'statut': statut,
            'prefix': prefix,
            'objectifs': ws_src.cell(row=r, column=3).value,
            'date_appro': _d(ws_src.cell(row=r, column=12).value),
            'date_accord': _d(ws_src.cell(row=r, column=13).value),
            'date_vigueur': _d(ws_src.cell(row=r, column=14).value),
            'date_demarrage': _d(ws_src.cell(row=r, column=15).value),
            'date_limite': _d(ws_src.cell(row=r, column=16).value),
            'mt_total': mt_total,
            'bailleur_mts': bailleur_mts,
            'taux_dec': taux_dec,
            'taux_phys': taux_phys,
            'observations': str(obs_val).strip() if obs_val else '',
        })

print(f"  {len(projets)} projets extraits (sections: En cours, Non demarres, En instance de cloture).")

# ── Génération du template d'import ──────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
EXAMPLE_FILL = PatternFill("solid", fgColor="F8FAFC")
FORMULA_FILL = PatternFill("solid", fgColor="FEFCE8")

def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30

def _auto_width(ws, ncols, min_w=12):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

def _write_cell(ws, row, col, value, fmt=None, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell

# ════════════════════ FEUILLE BAILLEURS ════════════════════
ws_b = wb.create_sheet("Bailleurs")
ws_b.sheet_properties.tabColor = "F59E0B"
ws_b.cell(row=1, column=1, value="BAILLEURS").font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
ws_b.cell(row=3, column=1, value="Sigle *")
ws_b.cell(row=3, column=2, value="Nom complet *")
ws_b.cell(row=3, column=3, value="Type de bailleur")
ws_b.cell(row=3, column=4, value="Description")
_style_header(ws_b, 3, 4)
row_b = 4
for sigle, (s, nom_complet) in BAILLEUR_SIGLES.items():
    ws_b.cell(row=row_b, column=1, value=s)
    ws_b.cell(row=row_b, column=2, value=nom_complet)
    ws_b.cell(row=row_b, column=3, value="Multilatéral")
    for c in range(1, 5):
        ws_b.cell(row=row_b, column=c).border = THIN_BORDER
    row_b += 1
_auto_width(ws_b, 4)

# ════════════════════ FEUILLE PROJETS ════════════════════
ws_p = wb.create_sheet("Projets")
ws_p.sheet_properties.tabColor = "EF4444"
ws_p.cell(row=1, column=1, value="PROJETS").font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
ws_p.cell(row=2, column=1, value="Feuille Projets — sans données financières (montants dans 'Accord de Financement', taux dans 'Décaissements')")
headers_p = [
    "Code projet *", "Titre *", "Description",
    "Statut",
    "Secteur (libellé)", "Bailleur principal (sigle)",
    "Date de signature", "Date de début", "Date de fin prévue", "Date de fin effective",
    "Zone géographique", "Responsable", "Structure responsable",
    "Code programme"
]
for c, h in enumerate(headers_p, 1):
    ws_p.cell(row=4, column=c, value=h)
_style_header(ws_p, 4, len(headers_p))

STATUT_MAP = {
    'en_cours':       "En cours d'exécution",
    'identification': 'Approuvé non démarré',
    'cloture':        'En instance de clôture',
}

row_p = 5
for p in projets:
    statut_label = STATUT_MAP.get(p['statut'], p['statut'])
    _write_cell(ws_p, row_p, 1, p['code'])
    _write_cell(ws_p, row_p, 2, p['nom'])
    _write_cell(ws_p, row_p, 3, str(p['objectifs'] or '').strip()[:500])
    _write_cell(ws_p, row_p, 4, statut_label)
    _write_cell(ws_p, row_p, 5, p['secteur'])
    _write_cell(ws_p, row_p, 6, p['sigle'])
    _write_cell(ws_p, row_p, 7, _fmt_date(p['date_accord']), 'YYYY-MM-DD')
    _write_cell(ws_p, row_p, 8, _fmt_date(p['date_demarrage']), 'YYYY-MM-DD')
    _write_cell(ws_p, row_p, 9, _fmt_date(p['date_limite']), 'YYYY-MM-DD')
    _write_cell(ws_p, row_p, 10, '')
    _write_cell(ws_p, row_p, 11, '')
    _write_cell(ws_p, row_p, 12, '')
    _write_cell(ws_p, row_p, 13, '')
    _write_cell(ws_p, row_p, 14, '')
    row_p += 1
_auto_width(ws_p, len(headers_p))

# ════════════════════ FEUILLE ACCORD DE FINANCEMENT ════════════════════
ws_f = wb.create_sheet("Accord de Financement")
ws_f.sheet_properties.tabColor = "3B82F6"
ws_f.cell(row=1, column=1, value="ACCORD DE FINANCEMENT").font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
ws_f.cell(row=2, column=1, value="Montant total XOF calculé automatiquement. Tous les montants sources sont en FCFA.")
headers_f = [
    "Code projet *", "Sigle bailleur *",
    "Type de financement",
    "Devise", "Montant total *", "Montant total XOF",
    "Date d'accord", "Référence accord", "Observations"
]
for c, h in enumerate(headers_f, 1):
    ws_f.cell(row=4, column=c, value=h)
_style_header(ws_f, 4, len(headers_f))

row_f = 5
for p in projets:
    bm = p['bailleur_mts']
    bailleur_entries = [(s, mt) for s, mt in bm.items() if mt and mt > 0]
    if not bailleur_entries:
        # Montant global sous le sigle principal
        bailleur_entries = [(p['sigle'], p['mt_total'])]
    for sigle, mt in bailleur_entries:
        _write_cell(ws_f, row_f, 1, p['code'])
        _write_cell(ws_f, row_f, 2, sigle)
        _write_cell(ws_f, row_f, 3, 'Prêt concessionnel')
        _write_cell(ws_f, row_f, 4, 'XOF')
        _write_cell(ws_f, row_f, 5, mt, '#,##0.00')
        _write_cell(ws_f, row_f, 6, mt, '#,##0.00', FORMULA_FILL)  # XOF=même valeur car déjà FCFA
        _write_cell(ws_f, row_f, 7, _fmt_date(p['date_accord']), 'YYYY-MM-DD')
        _write_cell(ws_f, row_f, 8, '')
        _write_cell(ws_f, row_f, 9, '')
        row_f += 1
_auto_width(ws_f, len(headers_f))

# ════════════════════ FEUILLE DÉCAISSEMENTS ════════════════════
ws_d = wb.create_sheet("Décaissements")
ws_d.sheet_properties.tabColor = "22C55E"
ws_d.cell(row=1, column=1, value="DÉCAISSEMENTS").font = Font(name="Calibri", bold=True, size=14, color=ACCENT)
ws_d.cell(row=2, column=1, value=(
    "Montant décaissé = Taux décaissement × Montant total. "
    "Taux d'exécution physique depuis le fichier source."
))
headers_d = [
    "Code projet *", "Sigle bailleur *",
    "Devise", "Montant décaissé cumulé *", "Montant décaissé XOF",
    "Taux de décaissement",
    "Taux de décaissement annuel prévu",
    "Taux d'exécution physique",
    "Taux d'exécution physique annuel prévu",
    "Montant dans le circuit de validation",
    "Montant dans le circuit de validation XOF",
    "Date de mise à jour *", "Description"
]
for c, h in enumerate(headers_d, 1):
    ws_d.cell(row=4, column=c, value=h)
_style_header(ws_d, 4, len(headers_d))

today_str = date.today().isoformat()
row_d = 5
for p in projets:
    bm = p['bailleur_mts']
    bailleur_entries = [(s, mt) for s, mt in bm.items() if mt and mt > 0]
    if not bailleur_entries:
        bailleur_entries = [(p['sigle'], p['mt_total'] or 0)]

    taux_dec_val = p['taux_dec']   # décimal 0-1 ou None
    taux_phys_val = p['taux_phys']  # décimal 0-1 ou None

    for sigle, mt in bailleur_entries:
        # Calcul du montant décaissé
        if mt and taux_dec_val is not None:
            mt_decaisse = round(mt * taux_dec_val, 2)
        else:
            mt_decaisse = None

        _write_cell(ws_d, row_d, 1, p['code'])
        _write_cell(ws_d, row_d, 2, sigle)
        _write_cell(ws_d, row_d, 3, 'XOF')
        _write_cell(ws_d, row_d, 4, mt_decaisse, '#,##0.00')
        _write_cell(ws_d, row_d, 5, mt_decaisse, '#,##0.00', FORMULA_FILL)  # XOF=même
        _write_cell(ws_d, row_d, 6, taux_dec_val, '0.00%')  # Taux décaissement
        _write_cell(ws_d, row_d, 7, None, '0.00%')          # Taux prévu annuel
        _write_cell(ws_d, row_d, 8, taux_phys_val, '0.00%') # Taux exec physique
        _write_cell(ws_d, row_d, 9, None, '0.00%')          # Taux exec physique prévu
        _write_cell(ws_d, row_d, 10, None, '#,##0.00')      # Montant circuit validation
        _write_cell(ws_d, row_d, 11, None, '#,##0.00', FORMULA_FILL)
        _write_cell(ws_d, row_d, 12, today_str, 'YYYY-MM-DD')  # Date MAJ
        _write_cell(ws_d, row_d, 13, p['observations'])
        row_d += 1
_auto_width(ws_d, len(headers_d))

# ════════════════════ SAUVEGARDE ════════════════════
wb.save(DST)
print(f"\nFichier genere : {DST}")
print(f"  Bailleurs   : {len(BAILLEUR_SIGLES)} lignes")
print(f"  Projets     : {len(projets)} lignes")
print(f"  Financements: {row_f - 5} lignes")
print(f"  Decaissements:{row_d - 5} lignes")
