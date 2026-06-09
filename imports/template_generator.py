"""
Génère le fichier Excel type à transmettre aux bailleurs.
4 feuilles : Bailleurs, Projets, Financements, Décaissements
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO


# ── Couleurs & Styles ──────────────────────────────────────────────
ORANGE = "F77F00"
GREEN = "009A44"
DARK = "1E293B"
LIGHT_ORANGE = "FFF7ED"
LIGHT_GREEN = "F0FDF4"
LIGHT_BLUE = "EFF6FF"
LIGHT_GRAY = "F8FAFC"
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=ORANGE)
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="64748B")

KEY_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
KEY_FONT = Font(name="Calibri", bold=True, size=11, color="C2410C")

EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="94A3B8")
EXAMPLE_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_key_column(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.fill = KEY_FILL
    cell.font = KEY_FONT


def _add_example_row(ws, row, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER


def _format_col(ws, col, fmt, row_start=5, row_end=500):
    """Applique un format numérique/date à une plage de cellules d'une colonne."""
    for r in range(row_start, row_end + 1):
        ws.cell(row=r, column=col).number_format = fmt


FORMULA_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
FORMULA_FONT = Font(name="Calibri", size=10, italic=True, color="92400E")


def _set_xof_formula(ws, col, devise_col, montant_col, taux_col=None, row_start=5, row_end=200):
    """
    Insère une formule de conversion en XOF dans col.
    Si taux_col est fourni : priorité au taux manuel, sinon conversion selon Devise.
    """
    D = get_column_letter(devise_col)
    M = get_column_letter(montant_col)
    for r in range(row_start, row_end + 1):
        base = (
            f'IF({D}{r}="XOF",{M}{r},IF({D}{r}="UC",{M}{r}*769.083,'
            f'IF({D}{r}="USD",{M}{r}*615,IF({D}{r}="EUR",{M}{r}*655.957,'
            f'IF({D}{r}="GBP",{M}{r}*780,{M}{r})))))'
        )
        if taux_col:
            T = get_column_letter(taux_col)
            formula = f'=IF({M}{r}="","",IF(AND({T}{r}<>"",{T}{r}>0),{M}{r}*{T}{r},{base}))'
        else:
            formula = f'=IF({M}{r}="","",{base})'
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = '#,##0.00'
        cell.fill = FORMULA_FILL
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER


def _auto_width(ws, col_count, min_width=14, max_width=40):
    for c in range(1, col_count + 1):
        letter = get_column_letter(c)
        lengths = []
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    lengths.append(len(str(cell.value)))
        best = max(lengths) if lengths else min_width
        ws.column_dimensions[letter].width = min(max(best + 3, min_width), max_width)


SECTEURS_CI = [
    'Agriculture/Développement rural',
    'Développement Urbain',
    'Eau & Assainissement',
    'Environnement & Climat',
    'Gouvernance',
    'Infrastructure & Transport',
    'Protection Sociale',
    'Santé',
    'Éducation/Formation',
    'Énergie',
    'Logement',
    'Finance Publique',
    'Autre',
]


def generate_template():
    """Retourne un BytesIO contenant le fichier Excel template."""
    from projets.models import Secteur as SecteurModel
    db_secteurs = list(SecteurModel.objects.order_by('nom').values_list('nom', flat=True))
    secteurs_list = db_secteurs if db_secteurs else SECTEURS_CI
    ex_secteur1 = secteurs_list[0] if len(secteurs_list) > 0 else "Santé"
    ex_secteur2 = secteurs_list[1] if len(secteurs_list) > 1 else "Éducation"

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 1 : BAILLEURS
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Bailleurs"
    ws1.sheet_properties.tabColor = ORANGE

    ws1.cell(row=1, column=1, value="FICHE BAILLEURS").font = TITLE_FONT
    ws1.cell(row=2, column=1, value="La colonne 'Sigle' sert de clé unique. Si un bailleur avec ce sigle existe déjà, ses informations seront mises à jour.").font = SUBTITLE_FONT

    headers_b = [
        "Sigle *", "Nom complet *", "Type de bailleur *",
        "Catégorie institutionnelle", "Pays du siège",
        "Description", "Site web", "Email de contact"
    ]
    for c, h in enumerate(headers_b, 1):
        ws1.cell(row=4, column=c, value=h)
    _style_header(ws1, 4, len(headers_b))

    # Exemple
    _add_example_row(ws1, 5, [
        "BM", "Banque Mondiale", "Multilatéral",
        "Institutions de Bretton Woods", "États-Unis",
        "Institution financière internationale", "https://worldbank.org", "info@worldbank.org"
    ])
    _add_example_row(ws1, 6, [
        "AFD", "Agence Française de Développement", "Bilatéral",
        "Coopération bilatérale", "France",
        "Agence de développement française", "https://afd.fr", ""
    ])

    # Validations
    type_bailleur_vals = '"Multilatéral,Bilatéral,Régional,Privé,ONG Internationale,Autre"'
    dv_type = DataValidation(type="list", formula1=type_bailleur_vals, allow_blank=False)
    dv_type.error = "Choisir : Multilatéral, Bilatéral, Régional, Privé, ONG Internationale, Autre"
    dv_type.errorTitle = "Type invalide"
    ws1.add_data_validation(dv_type)
    dv_type.add(f"C5:C1000")

    cat_vals = '"Institutions de Bretton Woods,Système des Nations Unies,Banques multilatérales de développement,Agence publique de développement,Institutions régionales africaines,Fonds de développement,Secteur privé / Fondations,ONG internationales,Autre"'
    dv_cat = DataValidation(type="list", formula1=cat_vals, allow_blank=True)
    ws1.add_data_validation(dv_cat)
    dv_cat.add(f"D5:D1000")

    _auto_width(ws1, len(headers_b))

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 2 : PROJETS
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Projets")
    ws2.sheet_properties.tabColor = GREEN

    ws2.cell(row=1, column=1, value="FICHE PROJETS").font = TITLE_FONT
    ws2.cell(row=2, column=1, value="La colonne 'Code projet' sert de clé unique. Si un projet avec ce code existe déjà, il sera mis à jour.").font = SUBTITLE_FONT

    # Colonnes Projets : SANS données financières (montants et taux dans les feuilles dédiées)
    headers_p = [
        "Code projet *", "Titre *", "Description",
        "Statut",
        "Secteur (libellé)", "Bailleur principal (sigle)",
        "Date de signature", "Date de début", "Date de fin prévue", "Date de fin effective",
        "Zone géographique", "Responsable", "Structure responsable",
        "Code programme",
        "Montant total projet XOF",
        "Part État (%)", "Part État",
    ]
    for c, h in enumerate(headers_p, 1):
        ws2.cell(row=4, column=c, value=h)
    _style_header(ws2, 4, len(headers_p))

    # Mettre en évidence les colonnes Montant total + Part État
    ETAT_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    ETAT_FONT = Font(name="Calibri", bold=True, size=11, color="166534")
    MT_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    MT_FONT = Font(name="Calibri", bold=True, size=11, color="1E40AF")
    # col 15 = Montant total projet XOF → style bleu
    cell_mt = ws2.cell(row=4, column=15)
    cell_mt.fill = MT_FILL
    cell_mt.font = MT_FONT
    cell_mt.alignment = HEADER_ALIGN
    cell_mt.border = THIN_BORDER
    # col 16, 17 = Part État (%, montant) → style vert
    for c_etat in [16, 17]:
        cell = ws2.cell(row=4, column=c_etat)
        cell.fill = ETAT_FILL
        cell.font = ETAT_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    _add_example_row(ws2, 5, [
        "PRJ-001", "Projet d'appui au secteur de la santé",
        "Projet visant à renforcer le système de santé",
        "En cours d'exécution",
        ex_secteur1, "BM",
        "2023-01-15", "2023-03-01", "2027-12-31", "",
        "Abidjan", "Direction SAN", "Ministère de la Santé", "PROG-001",
        35000000, 20.00, 7000000
    ])
    _add_example_row(ws2, 6, [
        "PRJ-002", "Projet éducatif numérique",
        "Projet pilote d'éducation numérique",
        "Approuvé mais non démarré",
        ex_secteur2, "AFD",
        "2025-06-01", "2025-09-01", "2028-08-31", "",
        "Bouaké", "Direction EDU", "Ministère de l'Éducation", "",
        "", 15.00, ""
    ])

    # Validation secteur — liste dynamique depuis feuille cachée RefSecteurs
    if secteurs_list:
        n_sec = len(secteurs_list)
        dv_secteur = DataValidation(type="list", formula1=f"=RefSecteurs!$A$1:$A${n_sec}", allow_blank=True)
        dv_secteur.prompt = "Choisissez ou saisissez le libellé complet du secteur"
        dv_secteur.promptTitle = "Secteur"
        ws2.add_data_validation(dv_secteur)
        dv_secteur.add("E5:E1000")

    # Col D = Statut (avec nouveaux statuts)
    statut_vals = '"Identification,Préparation,Négociation,Approuvé mais non démarré,En cours d\'exécution,En instance de clôture,Suspendu,Clôturé,Annulé"'
    dv_statut = DataValidation(type="list", formula1=statut_vals, allow_blank=True)
    ws2.add_data_validation(dv_statut)
    dv_statut.add("D5:D1000")

    for date_col in [7, 8, 9, 10]:
        _format_col(ws2, date_col, 'YYYY-MM-DD')
    _format_col(ws2, 14, '@')             # Code programme
    _format_col(ws2, 15, '#,##0')          # Montant total projet XOF
    _format_col(ws2, 16, '0.00')           # Part État (%)
    _format_col(ws2, 17, '#,##0.00')       # Part État montant

    # Formule auto : Part État montant = Montant total XOF × Part État (%) / 100
    for r in range(5, 201):
        cell_total = ws2.cell(row=r, column=15)
        cell_total.number_format = '#,##0'
        cell_total.border = THIN_BORDER
        cell_pct = ws2.cell(row=r, column=16)
        cell_pct.number_format = '0.00'
        cell_pct.border = THIN_BORDER
        cell_montant_etat = ws2.cell(row=r, column=17)
        cell_montant_etat.number_format = '#,##0.00'
        cell_montant_etat.border = THIN_BORDER
        # Formule : si Montant total (O) et % (P) sont renseignés → calculer Part État
        cell_montant_etat.value = f'=IF(AND(O{r}<>"",P{r}<>""),O{r}*P{r}/100,"")'

    _auto_width(ws2, len(headers_p))

    # ════════════════════════════════════════════════════════════════
    # FEUILLE CACHÉE : RÉFÉRENCE SECTEURS
    # ════════════════════════════════════════════════════════════════
    ws_ref = wb.create_sheet("RefSecteurs")
    ws_ref.sheet_state = "hidden"
    ws_ref.cell(row=1, column=1, value="Secteurs disponibles").font = Font(name="Calibri", bold=True, size=10)
    for idx, nom_sec in enumerate(secteurs_list, 1):
        ws_ref.cell(row=idx, column=1, value=nom_sec)
    ws_ref.column_dimensions["A"].width = 40

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 3 : PROGRAMMES
    # ════════════════════════════════════════════════════════════════
    ws_prog = wb.create_sheet("Programmes")
    ws_prog.sheet_properties.tabColor = "8B5CF6"

    ws_prog.cell(row=1, column=1, value="FICHE PROGRAMMES").font = TITLE_FONT
    ws_prog.cell(row=2, column=1, value=(
        "La colonne 'Code programme' est la clé unique. "
        "Les projets rattachés à un programme doivent y référencer son code."
    )).font = SUBTITLE_FONT

    headers_prog = [
        "Code programme *", "Nom *", "Description",
        "Statut",
        "Secteur (libellé)", "Bailleur principal (sigle)",
        "Date de signature", "Date de début", "Date de fin prévue", "Date de fin effective",
        "Zone géographique", "Responsable", "Structure responsable",
        "Objectif stratégique",
        "Montant total programme XOF",
        "Part État (%)", "Part État",
    ]
    for c, h in enumerate(headers_prog, 1):
        ws_prog.cell(row=4, column=c, value=h)
    _style_header(ws_prog, 4, len(headers_prog))

    _add_example_row(ws_prog, 5, [
        "PROG-001",
        "Programme de développement rural intégré",
        "Programme visant le renforcement des filières agricoles",
        "En cours d'exécution",
        ex_secteur1, "BM",
        "2023-01-15", "2024-01-01", "2028-12-31", "",
        "National", "Direction Développement Rural", "Ministère de l'Agriculture",
        "Améliorer la sécurité alimentaire et les revenus ruraux",
        80000000, 20.00, 16000000
    ])
    _add_example_row(ws_prog, 6, [
        "PROG-002",
        "Programme d'appui à l'éducation de base",
        "Amélioration de l'accès et de la qualité de l'éducation primaire",
        "En cours d'exécution",
        ex_secteur2, "AFD",
        "2023-03-01", "2023-06-01", "2027-05-31", "",
        "Abidjan, Bouaké", "Direction Éducation", "Ministère de l'Éducation",
        "Scolarisation universelle et amélioration des acquis",
        "", 15.00, ""
    ])

    if secteurs_list:
        n_sec = len(secteurs_list)
        dv_secteur_prog = DataValidation(
            type="list", formula1=f"=RefSecteurs!$A$1:$A${n_sec}", allow_blank=True
        )
        dv_secteur_prog.prompt = "Choisissez le libellé complet du secteur"
        dv_secteur_prog.promptTitle = "Secteur"
        ws_prog.add_data_validation(dv_secteur_prog)
        dv_secteur_prog.add("D5:D1000")

    statut_vals_prog = '"Identification,Préparation,Négociation,Approuvé mais non démarré,En cours d\'exécution,En instance de clôture,Suspendu,Clôturé,Annulé"'
    dv_statut_prog = DataValidation(type="list", formula1=statut_vals_prog, allow_blank=True)
    ws_prog.add_data_validation(dv_statut_prog)
    dv_statut_prog.add("D5:D1000")

    for date_col in [7, 8, 9, 10]:
        _format_col(ws_prog, date_col, 'YYYY-MM-DD')

    # Style colonnes financières
    MT_FILL_P = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    MT_FONT_P = Font(name="Calibri", bold=True, size=11, color="1E40AF")
    ETAT_FILL_P = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    ETAT_FONT_P = Font(name="Calibri", bold=True, size=11, color="166534")
    # col 15 = Montant total XOF → bleu
    hdr15 = ws_prog.cell(row=4, column=15)
    hdr15.fill = MT_FILL_P; hdr15.font = MT_FONT_P
    hdr15.alignment = HEADER_ALIGN; hdr15.border = THIN_BORDER
    # col 16, 17 = Part État → vert
    for c_e in [16, 17]:
        hdr = ws_prog.cell(row=4, column=c_e)
        hdr.fill = ETAT_FILL_P; hdr.font = ETAT_FONT_P
        hdr.alignment = HEADER_ALIGN; hdr.border = THIN_BORDER

    _format_col(ws_prog, 15, '#,##0')      # Montant total XOF
    _format_col(ws_prog, 16, '0.00')       # Part État (%)
    _format_col(ws_prog, 17, '#,##0.00')   # Part État montant

    # Formule auto : Part État montant = Montant total × % / 100
    for r in range(5, 201):
        cell_total = ws_prog.cell(row=r, column=15)
        cell_total.number_format = '#,##0'
        cell_total.border = THIN_BORDER
        cell_pct = ws_prog.cell(row=r, column=16)
        cell_pct.number_format = '0.00'
        cell_pct.border = THIN_BORDER
        cell_pe = ws_prog.cell(row=r, column=17)
        cell_pe.number_format = '#,##0.00'
        cell_pe.border = THIN_BORDER
        cell_pe.value = f'=IF(AND(O{r}<>"",P{r}<>""),O{r}*P{r}/100,"")'

    _auto_width(ws_prog, len(headers_prog))

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 4 : ACCORD DE FINANCEMENT
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Accord de Financement")
    ws3.sheet_properties.tabColor = "3B82F6"

    ws3.cell(row=1, column=1, value="ACCORD DE FINANCEMENT").font = TITLE_FONT
    ws3.cell(row=2, column=1, value=(
        "Code projet ou programme + Sigle bailleur + Type identifient un accord. "
        "'Montant total XOF' est calculé automatiquement selon la devise. "
        "Le montant dans le circuit de validation est saisi dans la feuille Décaissements."
    )).font = SUBTITLE_FONT

    headers_f = [
        "Code projet *", "Sigle bailleur *",
        "Type de financement",
        "Devise", "Montant total *", "Montant total XOF",
        "Date d'accord", "Référence accord", "Observations"
    ]
    for c, h in enumerate(headers_f, 1):
        ws3.cell(row=4, column=c, value=h)
    _style_header(ws3, 4, len(headers_f))

    _add_example_row(ws3, 5, [
        "PRJ-001", "BM", "Don", "USD", 35000000, None, "2023-01-15", "IDA-12345", "Financement principal"
    ])
    _add_example_row(ws3, 6, [
        "PRJ-001", "AFD", "Prêt concessionnel", "EUR", 8000000, None, "2023-06-01", "AFD-PRET-456", "Cofinancement 2ème bailleur"
    ])
    _add_example_row(ws3, 7, [
        "PROG-001", "BAD", "Prêt concessionnel", "UC", 10000000, None, "2023-01-01", "BAD-PROG-001", "Financement du programme"
    ])

    dv_type_fin = DataValidation(
        type="list",
        formula1='"Don,Prêt concessionnel,Prêt non concessionnel,Assistance technique,Cofinancement,Contrepartie nationale,Autre"',
        allow_blank=True
    )
    ws3.add_data_validation(dv_type_fin)
    dv_type_fin.add("C5:C1000")

    dv_devise_f = DataValidation(type="list", formula1='"UC,USD,EUR,XOF,GBP,JPY,CHF"', allow_blank=True)
    ws3.add_data_validation(dv_devise_f)
    dv_devise_f.add("D5:D1000")

    # Col E = Montant total, Col F = Montant total XOF (formule Devise × Montant)
    _format_col(ws3, 5, '#,##0.00')       # Montant total
    _set_xof_formula(ws3, col=6, devise_col=4, montant_col=5)
    _format_col(ws3, 7, 'YYYY-MM-DD')    # Date d'accord

    _auto_width(ws3, len(headers_f))

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 4 : DÉCAISSEMENTS
    # ════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Décaissements")
    ws4.sheet_properties.tabColor = "22C55E"

    ws4.cell(row=1, column=1, value="FICHE DÉCAISSEMENTS").font = TITLE_FONT
    ws4.cell(row=2, column=1, value="IMPORTANT : saisissez le montant CUMULÉ décaissé à la date de mise à jour (total depuis le début du projet). L'application calculera automatiquement la progression par différence.").font = SUBTITLE_FONT

    headers_d = [
        "Code projet *", "Sigle bailleur *",
        "Devise", "Montant décaissé cumulé *", "Montant décaissé XOF",
        "Taux de décaissement",
        "Taux de décaissement annuel prévu",
        "Taux d'exécution physique",
        "Taux d'exécution physique annuel prévu",
        "Montant dans le circuit de validation",
        "Montant dans le circuit de validation XOF",
        "Date de mise à jour *", "Commentaire / Justificatif"
    ]
    for c, h in enumerate(headers_d, 1):
        ws4.cell(row=4, column=c, value=h)
    _style_header(ws4, 4, len(headers_d))

    _add_example_row(ws4, 5, [
        "PRJ-001", "BM",
        "USD", 8000000, None, None, 0.25, 0.62, 0.55, 1500000, None,
        "2024-03-31", "Cumul au 31 mars 2024 (T1)"
    ])
    _add_example_row(ws4, 6, [
        "PRJ-001", "AFD",
        "EUR", 3500000, None, None, 0.20, 0.48, 0.40, 500000, None,
        "2024-03-31", "Cumul au 31 mars 2024 (T1)"
    ])

    # Validation Devise (col C = 3)
    dv_devise_d = DataValidation(type="list", formula1='"UC,USD,EUR,XOF,GBP,JPY,CHF"', allow_blank=True)
    ws4.add_data_validation(dv_devise_d)
    dv_devise_d.add("C5:C1000")

    # Préformatage
    _format_col(ws4, 4, '#,##0.00')   # Montant décaissé cumulé
    # Col E = Montant décaissé XOF
    _set_xof_formula(ws4, col=5, devise_col=3, montant_col=4)
    # Col F = Taux de décaissement (formule sans apostrophes dans le nom de feuille)
    SHEET_AF = "'Accord de Financement'"
    for r in range(5, 201):
        accord_ref = f"{SHEET_AF}!$A:$F"
        f = f'=IFERROR(IF(D{r}=0,"",E{r}/VLOOKUP(A{r},{accord_ref},6,0)),"")'
        cell = ws4.cell(row=r, column=6, value=f)
        cell.number_format = '0.00%'
        cell.fill = FORMULA_FILL
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER
    _format_col(ws4, 7, '0.00%')    # Taux déc annuel prévu
    _format_col(ws4, 8, '0.00%')    # Taux exécution physique
    _format_col(ws4, 9, '0.00%')    # Taux exécution physique prévu
    _format_col(ws4, 10, '#,##0.00') # Montant circuit validation
    # Col K = Montant circuit validation XOF
    _set_xof_formula(ws4, col=11, devise_col=3, montant_col=10)
    _format_col(ws4, 12, 'YYYY-MM-DD')  # Date de mise à jour

    _auto_width(ws4, len(headers_d))

    # ════════════════════════════════════════════════════════════════
    # FEUILLE 5 : INSTRUCTIONS
    # ════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Instructions")
    ws5.sheet_properties.tabColor = "8B5CF6"

    instructions = [
        ("GUIDE DE REMPLISSAGE — TEMPLATE SUIVI PROJETS BAILLEURS", TITLE_FONT),
        ("", None),
        ("STRUCTURE DU FICHIER (6 feuilles)", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("• Bailleurs : Organismes financeurs — clé unique = SIGLE.", None),
        ("• Projets : Données descriptives et administratives des projets (SANS montants financiers).", None),
        ("• Programmes : Programmes regroupant plusieurs projets (SANS montants financiers).", None),
        ("• Accord de Financement : Montants des accords de financement par bailleur et par projet/programme.", None),
        ("• Décaissements : État financier cumulé — montants décaissés, taux d'exécution, circuit de validation.", None),
        ("• Instructions : Ce guide.", None),
        ("", None),
        ("RÈGLES GÉNÉRALES", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("• Les colonnes marquées (*) sont OBLIGATOIRES.", None),
        ("• Supprimez les lignes d'exemple (fond gris) avant l'import.", None),
        ("• Format date : AAAA-MM-JJ (ex. : 2025-03-15).", None),
        ("• Montants : nombres bruts, sans espaces ni symboles (ex. : 64229926).", None),
        ("• Taux : décimal entre 0 et 1 (ex. : 0.45 = 45 %). Les colonnes formatées en % gèrent cela automatiquement.", None),
        ("• Utilisez les listes déroulantes pour Devise, Statut et Type de financement.", None),
        ("", None),
        ("FEUILLE ACCORD DE FINANCEMENT", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("• Ajoutez UNE LIGNE par bailleur et par projet/programme (cofinancement = plusieurs lignes).", None),
        ("• Codes programmes (PRG-...) et codes projets (PRJ-...) peuvent coexister dans cette feuille.", None),
        ("• 'Montant total XOF' : colonne à fond jaune = formule automatique, NE PAS modifier.", None),
        ("• Taux de conversion : 1 UC = 769,083 XOF | 1 USD = 615 XOF | 1 EUR = 655,957 XOF | 1 GBP = 780 XOF.", None),
        ("", None),
        ("FEUILLE DÉCAISSEMENTS", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("• UNE LIGNE par projet/bailleur avec l'état cumulé à la date de mise à jour.", None),
        ("• 'Montant décaissé cumulé' : total décaissé depuis le début du projet à cette date.", None),
        ("• 'Montant décaissé XOF' et 'Circuit de validation XOF' : colonnes à fond jaune = formules auto.", None),
        ("• 'Taux de décaissement' : formule auto = Décaissé XOF ÷ Total accord XOF (feuille Accord de Financement).", None),
        ("• Colonnes à saisir manuellement : Taux décaissement annuel prévu, Taux d'exécution physique (réel et prévu), Montant dans le circuit de validation.", None),
        ("", None),
        ("FEUILLE PROJETS — PART ÉTAT", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("• Part État (%) : pourcentage de la contrepartie nationale sur le montant total (ex. : 20 pour 20 %).", None),
        ("• Part État : montant brut de la contrepartie dans la devise du projet (calculé auto si % renseigné dans l'appli).", None),
        ("• Renseigner l'un ou l'autre — l'application calculera le champ manquant automatiquement.", None),
        ("", None),
        ("ORDRE DE REMPLISSAGE CONSEILLÉ", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("1. Bailleurs", None),
        ("2. Programmes", None),
        ("3. Projets (avec Part État si applicable)", None),
        ("4. Accord de Financement (montants des accords)", None),
        ("5. Décaissements (état cumulé financier)", None),
        ("", None),
        ("VALEURS ACCEPTÉES", Font(name="Calibri", bold=True, size=12, color=DARK)),
        ("Devise : UC, USD, EUR, XOF, GBP, JPY, CHF", None),
        ("Statut : Identification, Préparation, Négociation, Approuvé mais non démarré (ND), En cours d'exécution (EC), En instance de clôture (CL), Suspendu, Clôturé, Annulé", None),
        ("Type de financement : Don, Prêt concessionnel, Prêt non concessionnel, Assistance technique, Cofinancement, Contrepartie nationale, Autre", None),
        ("Secteur : libellé complet depuis la feuille RefSecteurs (liste déroulante disponible).", None),
    ]

    for i, (text, font) in enumerate(instructions, 1):
        cell = ws5.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        else:
            cell.font = Font(name="Calibri", size=10, color="334155")

    ws5.column_dimensions["A"].width = 100

    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
