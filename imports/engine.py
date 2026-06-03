"""
Moteur d'import intelligent : parse un Excel, détecte créations/mises à jour,
et exécute l'import avec un rapport détaillé.
"""
import openpyxl
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from collections import OrderedDict

from django.db.models import Sum
from bailleurs.models import Bailleur
from projets.models import Secteur, Projet, Programme, CommentaireProjet
from financements.models import Financement, Decaissement


# ── Mappings label → valeur DB ─────────────────────────────────────
TYPE_BAILLEUR_MAP = {
    'multilatéral': 'multilateral', 'multilateral': 'multilateral',
    'bilatéral': 'bilateral', 'bilateral': 'bilateral',
    'régional': 'regional', 'regional': 'regional',
    'privé': 'prive', 'prive': 'prive',
    'ong internationale': 'ong',
    'autre': 'autre',
}

CATEGORIE_MAP = {
    'institutions de bretton woods': 'bretton_woods',
    'système des nations unies': 'systeme_nu',
    'banques multilatérales de développement': 'banque_multilaterale',
    'agence publique de développement': 'agence_publique',
    'institutions régionales africaines': 'institution_regionale',
    'fonds de développement': 'fonds_developpement',
    'secteur privé / fondations': 'secteur_prive',
    'ong internationales': 'ong_internationale',
    'autre': 'autre',
    # anciens libellés
    'coopération bilatérale': 'cooperation_bilaterale',
    'fonds verticaux / thématiques': 'fonds_vertical',
}

STATUT_MAP = {
    'identification': 'identification',
    'préparation': 'preparation', 'preparation': 'preparation',
    'négociation': 'negociation', 'negociation': 'negociation',
    "en cours d'exécution": 'en_cours', 'en cours': 'en_cours', 'en_cours': 'en_cours',
    'suspendu': 'suspendu',
    'clôturé': 'cloture', 'cloture': 'cloture',
    'annulé': 'annule', 'annule': 'annule',
}

TYPE_FIN_MAP = {
    'don': 'don',
    'prêt concessionnel': 'pret_concessionnel', 'pret concessionnel': 'pret_concessionnel',
    'prêt non concessionnel': 'pret_non_concessionnel', 'pret non concessionnel': 'pret_non_concessionnel',
    'assistance technique': 'assistance_technique',
    'cofinancement': 'cofinancement',
    'contrepartie nationale': 'contrepartie',
    'autre': 'autre',
}


def _clean(val):
    """Nettoie une valeur de cellule."""
    if val is None:
        return ''
    if isinstance(val, str):
        return val.strip()
    return val


def _to_str(val):
    v = _clean(val)
    return str(v) if v != '' else ''


def _to_decimal(val):
    v = _clean(val)
    if v == '' or v is None:
        return Decimal('0')
    try:
        return Decimal(str(v).replace(',', '.').replace(' ', ''))
    except (InvalidOperation, ValueError):
        return None


def _to_date(val):
    v = _clean(val)
    if not v:
        return None
    if isinstance(v, (date, datetime)):
        return v if isinstance(v, date) and not isinstance(v, datetime) else v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        try:
            return datetime.strptime(str(v)[:10], '%d/%m/%Y').date()
        except (ValueError, TypeError):
            return None


def _build_devise_map(wb):
    """
    Construit {(code_lower, sigle_bailleur_lower): devise}
    depuis 'Accord de Financement' (nouveau nom) ou 'Financements' (ancien nom, fallback).
    """
    dmap = {}
    sheet_name = 'Accord de Financement' if 'Accord de Financement' in wb.sheetnames else 'Financements'
    if sheet_name in wb.sheetnames:
        for row in _read_sheet(wb[sheet_name]):
            code  = _to_str(row.get('Code projet', '')).strip().lower()
            sigle = _to_str(row.get('Sigle bailleur', '')).strip().lower()
            devise = (_to_str(row.get('Devise', '')) or 'XOF').upper().strip()
            if code and sigle:
                dmap[(code, sigle)] = devise
    return dmap


# Taux de conversion vers XOF (synchronisé avec projets.models.TAUX_VERS_FCFA)
_TAUX_VERS_XOF = {
    'XOF': Decimal('1'),
    'USD': Decimal('615.00'),
    'EUR': Decimal('655.957'),
    'GBP': Decimal('775.00'),
    'JPY': Decimal('4.10'),
    'CHF': Decimal('685.00'),
    'CNY': Decimal('85.00'),
    'UC':  Decimal('769.083'),
}


def _apply_uc_conversion(montant, code_projet, sigle_bailleur, devise_map):
    """Compatibilité ascendante : convertit vers XOF selon la devise du financement."""
    if devise_map is None:
        return montant
    key = (code_projet.strip().lower(), sigle_bailleur.strip().lower())
    devise = devise_map.get(key, 'XOF').upper()
    return _montant_to_xof(montant, devise)


def _montant_to_xof(montant, devise):
    """Convertit n'importe quelle devise vers XOF. Retourne le montant inchangé si XOF."""
    if montant is None or montant == 0:
        return montant
    devise = str(devise).upper().strip()
    taux = _TAUX_VERS_XOF.get(devise, Decimal('1'))
    if taux == Decimal('1'):
        return montant
    return (Decimal(str(montant)) * taux).quantize(Decimal('0.01'))


def _normalize_pct(val):
    """
    Normalise un taux en pourcentage.
    Si la valeur est entre 0 et 1 (exclus), considérée comme fraction Excel (0.74 → 74).
    Si la valeur est entre 1 et 100, gardée telle quelle.
    """
    if val is None:
        return Decimal('0')
    try:
        f = float(val)
    except (TypeError, ValueError):
        return Decimal('0')
    if 0 < f < 1:
        return Decimal(str(round(f * 100, 4)))
    if 0 <= f <= 100:
        return Decimal(str(f))
    return Decimal('0')


def _read_sheet(ws, header_row=4, data_start=5):
    """Lit une feuille à partir de header_row et retourne une liste de dicts."""
    rows_data = []
    headers = []
    for cell in ws[header_row]:
        h = _clean(cell.value)
        if h:
            headers.append(h.replace(' *', '').strip())
        else:
            headers.append(f'col_{cell.column}')

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        row_dict = {}
        for i, v in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = _clean(v)
        rows_data.append(row_dict)
    return rows_data


# ── ANALYSE (preview sans écriture) ───────────────────────────────

def _analyze_programmes(ws):
    rows = _read_sheet(ws)
    result = {'create': [], 'update': [], 'skip': [], 'errors': []}

    for i, row in enumerate(rows, 5):
        code = _to_str(row.get('Code programme', ''))
        nom = _to_str(row.get('Nom', ''))

        if not code:
            result['errors'].append(f"Ligne {i}: Code programme manquant")
            continue
        if not nom:
            result['errors'].append(f"Ligne {i}: Nom manquant")
            continue

        existing = Programme.objects.filter(code__iexact=code).first()
        info = f"[{code}] {nom}"
        if existing:
            result['update'].append({'line': i, 'info': info, 'detail': f"Mise \u00e0 jour de '{existing.nom}'"})
        else:
            result['create'].append({'line': i, 'info': info})

    return result


def analyze_file(file_obj):
    """
    Analyse un fichier Excel et retourne un rapport de preview.
    Retourne: {
        'bailleurs': {'create': [...], 'update': [...], 'skip': [...], 'errors': [...]},
        'programmes': {...},
        'projets': {...},
        'financements': {...},
        'decaissements': {...},
        'summary': {...}
    }
    """
    # Passe 1 : construire la map des devises avant toute lecture (read_only exhausts iterators)
    wb_tmp = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    devise_map = _build_devise_map(wb_tmp)
    wb_tmp.close()
    file_obj.seek(0)

    # Passe 2 : analyse complète
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    report = OrderedDict()

    # ── Bailleurs ──
    if 'Bailleurs' in wb.sheetnames:
        report['bailleurs'] = _analyze_bailleurs(wb['Bailleurs'])
    else:
        report['bailleurs'] = {'create': [], 'update': [], 'skip': [], 'errors': []}

    # ── Programmes ──
    if 'Programmes' in wb.sheetnames:
        report['programmes'] = _analyze_programmes(wb['Programmes'])
    else:
        report['programmes'] = {'create': [], 'update': [], 'skip': [], 'errors': []}

    # ── Projets ──
    if 'Projets' in wb.sheetnames:
        report['projets'] = _analyze_projets(wb['Projets'])
    else:
        report['projets'] = {'create': [], 'update': [], 'skip': [], 'errors': []}

    # ── Financements (Accord de Financement ou ancien nom) ──
    fin_ws_name = 'Accord de Financement' if 'Accord de Financement' in wb.sheetnames else 'Financements'
    if fin_ws_name in wb.sheetnames:
        report['financements'] = _analyze_financements(wb[fin_ws_name])
    else:
        report['financements'] = {'create': [], 'update': [], 'skip': [], 'errors': []}

    # ── Décaissements ──
    if 'Décaissements' in wb.sheetnames:
        report['decaissements'] = _analyze_decaissements(wb['Décaissements'], devise_map)
    else:
        report['decaissements'] = {'create': [], 'update': [], 'skip': [], 'errors': []}

    # Summary
    report['summary'] = {
        'total_create': sum(len(v['create']) for v in report.values() if isinstance(v, dict) and 'create' in v),
        'total_update': sum(len(v['update']) for v in report.values() if isinstance(v, dict) and 'update' in v),
        'total_errors': sum(len(v['errors']) for v in report.values() if isinstance(v, dict) and 'errors' in v),
    }

    wb.close()
    return report


def _analyze_bailleurs(ws):
    rows = _read_sheet(ws)
    result = {'create': [], 'update': [], 'skip': [], 'errors': []}

    for i, row in enumerate(rows, 5):
        sigle = _to_str(row.get('Sigle', ''))
        nom = _to_str(row.get('Nom complet', ''))

        if not sigle:
            result['errors'].append(f"Ligne {i}: Sigle manquant")
            continue
        if not nom:
            result['errors'].append(f"Ligne {i}: Nom complet manquant")
            continue

        existing = Bailleur.objects.filter(sigle__iexact=sigle).first()
        info = f"{sigle} - {nom}"
        if existing:
            result['update'].append({'line': i, 'info': info, 'detail': f"Mise à jour de '{existing.nom}'"})
        else:
            result['create'].append({'line': i, 'info': info})

    return result


def _analyze_projets(ws):
    rows = _read_sheet(ws)
    result = {'create': [], 'update': [], 'skip': [], 'errors': []}

    for i, row in enumerate(rows, 5):
        code = _to_str(row.get('Code projet', ''))
        titre = _to_str(row.get('Titre', ''))

        if not code:
            result['errors'].append(f"Ligne {i}: Code projet manquant")
            continue
        if not titre:
            result['errors'].append(f"Ligne {i}: Titre manquant")
            continue

        bailleur_sigle = _to_str(row.get('Bailleur principal (sigle)', ''))
        if bailleur_sigle and not Bailleur.objects.filter(sigle__iexact=bailleur_sigle).exists():
            result['errors'].append(f"Ligne {i}: Bailleur '{bailleur_sigle}' introuvable")
            continue

        secteur_val = _to_str(row.get('Secteur (libellé)', '') or row.get('Secteur (code)', ''))
        if secteur_val and not (
            Secteur.objects.filter(nom__iexact=secteur_val).exists() or
            Secteur.objects.filter(nom__icontains=secteur_val).exists() or
            Secteur.objects.filter(code__iexact=secteur_val).exists()
        ):
            result['errors'].append(f"Ligne {i}: Secteur '{secteur_val}' introuvable")
            continue

        existing = Projet.objects.filter(code__iexact=code).first()
        info = f"[{code}] {titre}"
        if existing:
            result['update'].append({'line': i, 'info': info, 'detail': f"Mise à jour de '{existing.titre}'"})
        else:
            result['create'].append({'line': i, 'info': info})

    return result


def _analyze_financements(ws):
    rows = _read_sheet(ws)
    result = {'create': [], 'update': [], 'skip': [], 'errors': []}

    for i, row in enumerate(rows, 5):
        code_projet = _to_str(row.get('Code projet', ''))
        sigle_bailleur = _to_str(row.get('Sigle bailleur', ''))

        if not code_projet:
            result['errors'].append(f"Ligne {i}: Code projet manquant")
            continue
        if not sigle_bailleur:
            result['errors'].append(f"Ligne {i}: Sigle bailleur manquant")
            continue

        projet = Projet.objects.filter(code__iexact=code_projet).first()
        bailleur = Bailleur.objects.filter(sigle__iexact=sigle_bailleur).first()
        if not bailleur:
            result['errors'].append(f"Ligne {i}: Bailleur '{sigle_bailleur}' introuvable")
            continue

        fin_devise = _to_str(row.get('Devise', 'UC')) or 'UC'
        montant_raw = _to_decimal(row.get('Montant total', '') or row.get('Montant total *', '')) or Decimal('0')
        montant_xof = _to_decimal(row.get('Montant total XOF', ''))
        montant = montant_xof if (montant_xof and montant_xof > 0) else _montant_to_xof(montant_raw, fin_devise)

        type_raw = _to_str(row.get('Type de financement', '')).lower()
        type_fin = TYPE_FIN_MAP.get(type_raw, 'don')
        info = f"{sigle_bailleur} → {code_projet} ({montant:,.0f} XOF)"

        # Vérifier si c'est un projet ou un programme
        if projet:
            existing = Financement.objects.filter(projet=projet, bailleur=bailleur, type_financement=type_fin).first()
            if existing:
                result['update'].append({'line': i, 'info': info, 'detail': f"Montant {existing.montant_engage:,.0f} → {montant:,.0f}"})
            else:
                result['create'].append({'line': i, 'info': info})
        else:
            programme = Programme.objects.filter(code__iexact=code_projet).first()
            if programme:
                result['update'].append({'line': i, 'info': f"Programme {info}", 'detail': "Mise à jour montant_total"})
            else:
                result['errors'].append(f"Ligne {i}: Code '{code_projet}' introuvable (ni projet ni programme)")

    return result


def _get_col_dec(row, *keys):
    """Lit la première clé non vide parmi les alternatives."""
    for k in keys:
        v = row.get(k)
        if v not in (None, ''):
            return v
    return ''


def _analyze_decaissements(ws, devise_map=None):
    """
    Analyse de la feuille Décaissements.
    Les montants sont CUMULÉS : on prévoit création ou mise à jour
    en comparant avec le cumul existant pour chaque financement.
    Clé : financement (code_projet + sigle_bailleur) + date_mise_a_jour.
    La conversion UC→XOF est appliquée si la devise source est UC (via devise_map).
    """
    rows = _read_sheet(ws)
    result = {'create': [], 'update': [], 'skip': [], 'errors': []}

    for i, row in enumerate(rows, 5):
        code_projet = _to_str(row.get('Code projet', ''))
        sigle_bailleur = _to_str(row.get('Sigle bailleur', ''))
        montant_cumule = _to_decimal(
            _get_col_dec(row, 'Montant décaissé cumulé', 'Montant décaissé')
        )
        date_maj = _to_date(
            _get_col_dec(row, 'Date de mise à jour', 'Date de décaissement')
        )
        # Conversion vers XOF : priorité colonne "Montant XOF", puis Taux inline, puis devise_map
        if montant_cumule and montant_cumule > 0:
            montant_xof_direct = _to_decimal(row.get('Montant XOF', ''))
            if montant_xof_direct and montant_xof_direct > 0:
                montant_cumule = montant_xof_direct
            else:
                dec_devise = _to_str(row.get('Devise', ''))
                dec_taux = _to_decimal(row.get('Taux de change vs XOF', 0)) or Decimal('0')
                if dec_taux > 0:
                    montant_cumule = (montant_cumule * dec_taux).quantize(Decimal('0.01'))
                elif dec_devise:
                    montant_cumule = _montant_to_xof(montant_cumule, dec_devise)
                else:
                    montant_cumule = _apply_uc_conversion(montant_cumule, code_projet, sigle_bailleur, devise_map)

        if not code_projet:
            result['errors'].append(f"Ligne {i}: Code projet manquant")
            continue
        if not montant_cumule or montant_cumule <= 0:
            result['skip'].append({'line': i, 'info': f"{code_projet} | aucun décaissement enregistré"})
            continue
        if not date_maj:
            result['errors'].append(f"Ligne {i}: Date de mise à jour manquante (obligatoire si montant > 0)")
            continue

        projet = Projet.objects.filter(code__iexact=code_projet).first()
        if not projet:
            result['errors'].append(f"Ligne {i}: Projet '{code_projet}' introuvable")
            continue

        bailleur = Bailleur.objects.filter(sigle__iexact=sigle_bailleur).first() if sigle_bailleur else None
        if not bailleur:
            result['errors'].append(f"Ligne {i}: Bailleur '{sigle_bailleur}' introuvable")
            continue

        financement = Financement.objects.filter(projet=projet, bailleur=bailleur).first()
        if not financement:
            result['errors'].append(f"Ligne {i}: Aucun financement {sigle_bailleur} → {code_projet}")
            continue

        existing = Decaissement.objects.filter(
            financement=financement, date_decaissement=date_maj
        ).first()

        info = f"{code_projet} | {sigle_bailleur} | cumulé {float(montant_cumule):,.0f} | {date_maj}"
        if existing:
            total_sans = float(
                Decaissement.objects.filter(financement=financement)
                .exclude(pk=existing.pk)
                .aggregate(t=Sum('montant'))['t'] or 0
            )
            new_amt = float(montant_cumule) - total_sans
            if abs(new_amt - float(existing.montant)) > 0.01:
                result['update'].append({'line': i, 'info': info,
                    'detail': f"δ {existing.montant:,.0f} → {new_amt:,.2f}"})
            else:
                result['skip'].append({'line': i, 'info': info + ' (inchangé)'})
        else:
            total_existant = float(
                Decaissement.objects.filter(financement=financement)
                .aggregate(t=Sum('montant'))['t'] or 0
            )
            delta = float(montant_cumule) - total_existant
            if delta > 0:
                result['create'].append({'line': i, 'info': info,
                    'detail': f"nouvel incrément +{delta:,.2f}"})
            else:
                result['skip'].append({'line': i, 'info': info + ' (cumul déjà intégré)'})

    return result


# ── EXÉCUTION (écriture en base) ──────────────────────────────────

def execute_import(file_obj):
    """
    Exécute l'import et retourne un rapport avec les compteurs.
    """
    # Passe 1 : lire la map des devises avant que les feuilles soient exhaustées
    wb_tmp = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    devise_map = _build_devise_map(wb_tmp)
    wb_tmp.close()
    file_obj.seek(0)

    # Passe 2 : import complet
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    counts = OrderedDict([
        ('bailleurs', {'created': 0, 'updated': 0, 'errors': []}),
        ('programmes', {'created': 0, 'updated': 0, 'errors': []}),
        ('projets', {'created': 0, 'updated': 0, 'errors': []}),
        ('financements', {'created': 0, 'updated': 0, 'errors': []}),
        ('decaissements', {'created': 0, 'updated': 0, 'errors': []}),
    ])

    # Import dans l'ordre: Bailleurs → Programmes → Projets → Financements → Décaissements
    if 'Bailleurs' in wb.sheetnames:
        _import_bailleurs(wb['Bailleurs'], counts['bailleurs'])

    if 'Programmes' in wb.sheetnames:
        _import_programmes(wb['Programmes'], counts['programmes'])

    if 'Projets' in wb.sheetnames:
        _import_projets(wb['Projets'], counts['projets'])

    fin_ws_name = 'Accord de Financement' if 'Accord de Financement' in wb.sheetnames else 'Financements'
    if fin_ws_name in wb.sheetnames:
        # Réinitialiser montant_total programmes avant recalcul
        Programme.objects.all().update(montant_total=Decimal('0'))
        _import_financements(wb[fin_ws_name], counts['financements'])
        # Mettre à jour montant_total projets depuis la somme de leurs financements
        from django.db.models import Sum as _Sum
        for p in Projet.objects.all():
            total = Financement.objects.filter(projet=p).aggregate(s=_Sum('montant_engage'))['s']
            if total is not None:
                p.montant_total = total
                p.devise = 'XOF'
                p.save(update_fields=['montant_total', 'devise'])

    if 'Décaissements' in wb.sheetnames:
        _import_decaissements(wb['Décaissements'], counts['decaissements'], devise_map)

    wb.close()
    return counts


def _import_programmes(ws, counts):
    rows = _read_sheet(ws)
    for i, row in enumerate(rows, 5):
        try:
            code = _to_str(row.get('Code programme', ''))
            nom = _to_str(row.get('Nom', ''))
            if not code or not nom:
                continue

            secteur_val = _to_str(row.get('Secteur (libellé)', '') or row.get('Secteur', ''))
            secteur = None
            if secteur_val:
                secteur = (
                    Secteur.objects.filter(nom__iexact=secteur_val).first() or
                    Secteur.objects.filter(nom__icontains=secteur_val).first()
                )
                if not secteur:
                    secteur, _ = Secteur.objects.get_or_create(
                        nom=secteur_val, defaults={'code': secteur_val[:20].upper().replace(' ', '_')}
                    )

            bailleur_sigle = _to_str(row.get('Bailleur principal (sigle)', ''))
            bailleur = Bailleur.objects.filter(sigle__iexact=bailleur_sigle).first() if bailleur_sigle else None

            statut_raw = _to_str(row.get('Statut', '')).lower()

            defaults = {
                'nom': nom,
                'description': _to_str(row.get('Description', '')),
                'secteur': secteur,
                'bailleur_principal': bailleur,
                'date_signature': _to_date(row.get('Date de signature', '')),
                'date_debut': _to_date(row.get('Date de début', '') or row.get('Date début', '')),
                'date_fin': _to_date(row.get('Date de fin prévue', '') or row.get('Date fin', '')),
                'date_fin_effective': _to_date(row.get('Date de fin effective', '')),
                'statut': STATUT_MAP.get(statut_raw, 'identification') if statut_raw else 'identification',
                'zone_geographique': _to_str(row.get('Zone géographique', '')),
                'responsable': _to_str(row.get('Responsable', '')),
                'structure_responsable': _to_str(row.get('Structure responsable', '')),
                'objectif_strategique': _to_str(row.get('Objectif stratégique', '') or row.get('Objectif strategique', '')),
            }

            existing = Programme.objects.filter(code__iexact=code).first()
            if existing:
                for k, v in defaults.items():
                    setattr(existing, k, v)
                existing.save()
                counts['updated'] += 1
            else:
                Programme.objects.create(code=code, **defaults)
                counts['created'] += 1

        except Exception as e:
            counts['errors'].append(f"Ligne {i}: {str(e)}")


def _import_bailleurs(ws, counts):
    rows = _read_sheet(ws)
    for i, row in enumerate(rows, 5):
        try:
            sigle = _to_str(row.get('Sigle', ''))
            nom = _to_str(row.get('Nom complet', ''))
            if not sigle or not nom:
                continue

            type_raw = _to_str(row.get('Type de bailleur', '')).lower()
            cat_raw = _to_str(row.get('Catégorie institutionnelle', '')).lower()

            defaults = {
                'nom': nom,
                'type_bailleur': TYPE_BAILLEUR_MAP.get(type_raw, 'autre'),
                'categorie_institutionnelle': CATEGORIE_MAP.get(cat_raw, ''),
                'pays_siege': _to_str(row.get('Pays du siège', '')),
                'description': _to_str(row.get('Description', '')),
                'site_web': _to_str(row.get('Site web', '')),
                'contact_email': _to_str(row.get('Email de contact', '')),
            }

            obj, created = Bailleur.objects.update_or_create(
                sigle__iexact=sigle,
                defaults={**defaults, 'sigle': sigle}
            )

            if created:
                counts['created'] += 1
            else:
                counts['updated'] += 1

        except Exception as e:
            counts['errors'].append(f"Ligne {i}: {str(e)}")


def _import_projets(ws, counts):
    rows = _read_sheet(ws)
    for i, row in enumerate(rows, 5):
        try:
            code = _to_str(row.get('Code projet', ''))
            titre = _to_str(row.get('Titre', ''))
            if not code or not titre:
                continue

            bailleur_sigle = _to_str(row.get('Bailleur principal (sigle)', ''))
            bailleur = Bailleur.objects.filter(sigle__iexact=bailleur_sigle).first() if bailleur_sigle else None

            secteur_val = _to_str(row.get('Secteur (libellé)', '') or row.get('Secteur (code)', ''))
            if secteur_val:
                secteur = (
                    Secteur.objects.filter(nom__iexact=secteur_val).first() or
                    Secteur.objects.filter(nom__icontains=secteur_val).first() or
                    Secteur.objects.filter(code__iexact=secteur_val).first()
                )
                if not secteur:
                    secteur, _ = Secteur.objects.get_or_create(
                        nom=secteur_val, defaults={'code': secteur_val[:20].upper().replace(' ', '_')}
                    )
            else:
                secteur = None

            statut_raw = _to_str(row.get('Statut', '')).lower()

            defaults = {
                'titre': titre,
                'description': _to_str(row.get('Description', '')),
                'secteur': secteur,
                'bailleur_principal': bailleur,
                'date_signature': _to_date(row.get('Date de signature', '')),
                'date_debut': _to_date(row.get('Date de début', '')),
                'date_fin_prevue': _to_date(row.get('Date de fin prévue', '')),
                'date_fin_effective': _to_date(row.get('Date de fin effective', '')),
                'structure_responsable': _to_str(row.get('Structure responsable', '')),
                'statut': STATUT_MAP.get(statut_raw, 'identification') if statut_raw else 'identification',
                'zone_geographique': _to_str(row.get('Zone géographique', '')),
                'responsable': _to_str(row.get('Responsable', '')),
            }

            prog_code = _to_str(row.get('Code programme', ''))
            if prog_code:
                prog = Programme.objects.filter(code__iexact=prog_code).first()
                if prog:
                    defaults['programme'] = prog

            obj, created = Projet.objects.update_or_create(
                code__iexact=code,
                defaults={**defaults, 'code': code}
            )

            if created:
                counts['created'] += 1
            else:
                counts['updated'] += 1

        except Exception as e:
            counts['errors'].append(f"Ligne {i}: {str(e)}")


def _import_financements(ws, counts):
    """
    Lit la feuille 'Accord de Financement' (ou ancienne 'Financements').
    - Codes PRJ-... : crée/met à jour un Financement lié au Projet.
    - Codes PRG-... : met à jour Programme.montant_total.
    """
    rows = _read_sheet(ws)
    for i, row in enumerate(rows, 5):
        try:
            code = _to_str(row.get('Code projet', ''))
            sigle_bailleur = _to_str(row.get('Sigle bailleur', ''))
            if not code or not sigle_bailleur:
                continue

            type_raw = _to_str(row.get('Type de financement', '')).lower()
            fin_devise = _to_str(row.get('Devise', 'UC')) or 'UC'
            montant_raw = _to_decimal(row.get('Montant total', '') or row.get('Montant total *', '')) or Decimal('0')
            # 'Montant total XOF' est calculé par formule Excel — lire en priorité
            montant_xof = _to_decimal(row.get('Montant total XOF', ''))
            if montant_xof and montant_xof > 0:
                montant = montant_xof
            else:
                montant = _montant_to_xof(montant_raw, fin_devise)

            bailleur = Bailleur.objects.filter(sigle__iexact=sigle_bailleur).first()
            if not bailleur:
                counts['errors'].append(f"Ligne {i}: Bailleur '{sigle_bailleur}' introuvable")
                continue

            # Essayer d'abord un Projet, puis un Programme
            projet = Projet.objects.filter(code__iexact=code).first()
            if projet:
                defaults = {
                    'montant_engage': montant,
                    'montant_circuit_validation': Decimal('0'),
                    'devise': 'XOF',
                    'date_accord': _to_date(row.get("Date d'accord", '')),
                    'reference': _to_str(row.get('Référence accord', '')),
                    'observations': _to_str(row.get('Observations', '')),
                }
                type_fin = TYPE_FIN_MAP.get(type_raw, 'don')
                obj, created = Financement.objects.update_or_create(
                    projet=projet, bailleur=bailleur, type_financement=type_fin,
                    defaults=defaults
                )
                if created:
                    counts['created'] += 1
                else:
                    counts['updated'] += 1
            else:
                programme = Programme.objects.filter(code__iexact=code).first()
                if programme:
                    # Additionner les montants si plusieurs bailleurs
                    programme.montant_total = (programme.montant_total or Decimal('0')) + montant
                    programme.devise = 'XOF'
                    programme.save(update_fields=['montant_total', 'devise'])
                    counts['updated'] += 1
                else:
                    counts['errors'].append(f"Ligne {i}: Code '{code}' introuvable (ni projet ni programme)")

        except Exception as e:
            counts['errors'].append(f"Ligne {i}: {str(e)}")


def _import_decaissements(ws, counts, devise_map=None):
    """
    Import avec logique cumulée → incrémentale.
    Nouvelles colonnes (template v2) :
      - Devise / Montant décaissé cumulé / Montant décaissé XOF (formule)
      - Taux de décaissement annuel prévu → projet.taux_decaissement_prevu_annee
      - Taux d'exécution physique         → projet.taux_avancement
      - Taux d'exécution physique annuel prévu → projet.taux_avancement_financier
      - Montant dans le circuit de validation / XOF → financement.montant_circuit_validation
      - Date de mise à jour / Description
    """
    rows = _read_sheet(ws)
    for i, row in enumerate(rows, 5):
        try:
            code_projet = _to_str(row.get('Code projet', ''))
            sigle_bailleur = _to_str(row.get('Sigle bailleur', ''))
            if not code_projet or not sigle_bailleur:
                continue

            # ── Montant décaissé en XOF ──────────────────────────────────
            # Priorité 1 : colonne "Montant décaissé XOF" (formule Excel)
            montant_xof_direct = _to_decimal(
                row.get('Montant décaissé XOF', '') or row.get('Montant XOF', '')
            )
            montant_cumule_raw = _to_decimal(
                _get_col_dec(row, 'Montant décaissé cumulé *', 'Montant décaissé cumulé', 'Montant décaissé')
            )
            if montant_xof_direct and montant_xof_direct > 0:
                montant_cumule = montant_xof_direct
            elif montant_cumule_raw and montant_cumule_raw > 0:
                dec_devise = _to_str(row.get('Devise', ''))
                if dec_devise:
                    montant_cumule = _montant_to_xof(montant_cumule_raw, dec_devise)
                else:
                    montant_cumule = _apply_uc_conversion(montant_cumule_raw, code_projet, sigle_bailleur, devise_map)
            else:
                continue  # pas de montant : ignorer la ligne

            # ── Autres colonnes ──────────────────────────────────────────
            date_maj = _to_date(
                _get_col_dec(row, 'Date de mise à jour *', 'Date de mise à jour', 'Date de décaissement')
            )
            description = _to_str(
                row.get('Commentaire / Justificatif', '')
                or row.get('Description', '')
            )

            taux_prevu = _normalize_pct(_to_decimal(
                row.get('Taux de décaissement annuel prévu', 0)
            ))
            taux_phys = _normalize_pct(_to_decimal(
                row.get("Taux d'exécution physique", 0)
            ))
            taux_phys_prevu = _normalize_pct(_to_decimal(
                row.get("Taux d'exécution physique annuel prévu", 0)
            ))

            # Montant circuit de validation en XOF
            circuit_xof = _to_decimal(
                row.get('Montant dans le circuit de validation XOF', '')
                or row.get('Montant circuit de validation XOF', '')
            )
            circuit_raw = _to_decimal(
                row.get('Montant dans le circuit de validation', 0)
                or row.get('Montant circuit de validation', 0)
            ) or Decimal('0')
            if circuit_xof and circuit_xof >= 0:
                circuit = circuit_xof
            elif circuit_raw > 0:
                dec_devise = _to_str(row.get('Devise', ''))
                circuit = _montant_to_xof(circuit_raw, dec_devise) if dec_devise else circuit_raw
            else:
                circuit = Decimal('0')

            if not date_maj:
                counts['errors'].append(f"Ligne {i}: Date de mise à jour manquante")
                continue

            projet = Projet.objects.filter(code__iexact=code_projet).first()
            bailleur = Bailleur.objects.filter(sigle__iexact=sigle_bailleur).first()
            if not projet or not bailleur:
                counts['errors'].append(f"Ligne {i}: Projet ou bailleur introuvable")
                continue

            financement = Financement.objects.filter(projet=projet, bailleur=bailleur).first()
            if not financement:
                counts['errors'].append(f"Ligne {i}: Aucun financement {sigle_bailleur} → {code_projet}")
                continue

            # ── Mettre à jour le circuit de validation sur le financement ──
            financement.montant_circuit_validation = circuit
            financement.save(update_fields=['montant_circuit_validation'])

            # ── Mettre à jour les taux sur le projet ──
            projet_fields = {}
            if taux_prevu is not None and taux_prevu >= 0:
                projet_fields['taux_decaissement_prevu_annee'] = taux_prevu
            if taux_phys is not None and taux_phys >= 0:
                projet_fields['taux_avancement'] = taux_phys
            if taux_phys_prevu is not None and taux_phys_prevu >= 0:
                projet_fields['taux_avancement_financier'] = taux_phys_prevu
            if projet_fields:
                for k, v in projet_fields.items():
                    setattr(projet, k, v)
                projet.save(update_fields=list(projet_fields.keys()))

            # ── Logique cumulée → incrémentale ──
            existing = Decaissement.objects.filter(
                financement=financement, date_decaissement=date_maj
            ).first()

            if existing:
                total_sans = float(
                    Decaissement.objects.filter(financement=financement)
                    .exclude(pk=existing.pk)
                    .aggregate(t=Sum('montant'))['t'] or 0
                )
                new_montant = round(float(montant_cumule) - total_sans, 2)
                if new_montant > 0 and abs(new_montant - float(existing.montant)) > 0.01:
                    existing.montant = Decimal(str(new_montant))
                    existing.description = description or existing.description
                    existing.save()
                    counts['updated'] += 1
            else:
                total_existant = float(
                    Decaissement.objects.filter(financement=financement)
                    .aggregate(t=Sum('montant'))['t'] or 0
                )
                delta = round(float(montant_cumule) - total_existant, 2)
                if delta > 0:
                    Decaissement.objects.create(
                        financement=financement,
                        montant=Decimal(str(delta)),
                        date_decaissement=date_maj,
                        reference='',
                        description=description,
                    )
                    counts['created'] += 1

            # ── Sauvegarder la description/observations comme CommentaireProjet ──
            obs_text = description.strip() if description else ''
            if obs_text:
                CommentaireProjet.objects.get_or_create(
                    projet=projet,
                    source='import_excel',
                    contenu=obs_text,
                    defaults={
                        'type_commentaire': 'observation',
                        'niveau': 'information',
                        'date_commentaire': date_maj,
                    },
                )

        except Exception as e:
            counts['errors'].append(f"Ligne {i}: {str(e)}")
